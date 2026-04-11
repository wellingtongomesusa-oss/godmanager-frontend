import os
import csv
import io
import json
import re
import uuid
from calendar import monthrange
from collections import Counter, defaultdict
from datetime import datetime, timedelta, date
from functools import wraps
from decimal import Decimal

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, Response, send_from_directory, current_app
from sqlalchemy import text
from flask_sqlalchemy import SQLAlchemy
from flask_session import Session
import bcrypt
from urllib.parse import urlparse
import base64

import requests as req
import hashlib
import hmac

# Initialize extensions
db = SQLAlchemy()
sess = Session()


def hash_password(raw_password: str) -> str:
    """Hash password using bcrypt."""
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(raw_password.encode("utf-8"), salt).decode("utf-8")


def is_password_strong(password: str) -> bool:
    """Validate password strength."""
    if len(password) < 8:
        return False
    has_upper = any(c.isupper() for c in password)
    has_lower = any(c.islower() for c in password)
    has_digit = any(c.isdigit() for c in password)
    has_special = any(not c.isalnum() for c in password)
    return has_upper and has_lower and has_digit and has_special


def generate_session_token() -> str:
    """Generate secure session token."""
    import secrets
    return secrets.token_urlsafe(32)


def login_required(f):
    """Decorator - login desativado, sempre permite acesso."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """Decorator - admin desativado, sempre permite acesso."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        return f(*args, **kwargs)
    return decorated_function


# Database Models
class User(db.Model):
    __tablename__ = "users"
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    first_name = db.Column(db.String(80))
    last_name = db.Column(db.String(80))
    is_admin = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    failed_login_attempts = db.Column(db.Integer, default=0)
    locked_until = db.Column(db.DateTime, nullable=True)
    email_verified = db.Column(db.Boolean, default=False)
    two_factor_enabled = db.Column(db.Boolean, default=False)
    two_factor_secret = db.Column(db.String(255), nullable=True)

    def __repr__(self):
        return f"<User {self.username}>"


class UserSession(db.Model):
    __tablename__ = "user_sessions"
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="CASCADE"), nullable=False)
    session_token = db.Column(db.String(255), unique=True, nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    expires_at = db.Column(db.DateTime, nullable=False, index=True)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.Text)
    is_active = db.Column(db.Boolean, default=True)

    user = db.relationship("User", backref="sessions")


class AccountRequest(db.Model):
    __tablename__ = "account_requests"
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), nullable=False)
    first_name = db.Column(db.String(80), nullable=False)
    last_name = db.Column(db.String(80), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    status = db.Column(db.String(20), default="pending", index=True)
    admin_notes = db.Column(db.Text)
    processed_by = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    processed_at = db.Column(db.DateTime)


class AuditLog(db.Model):
    __tablename__ = "audit_logs"
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="SET NULL"), nullable=True, index=True)
    action = db.Column(db.String(100), nullable=False)
    details = db.Column(db.Text)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class PhoneLine(db.Model):
    """Linhas telefônicas — modelo AT&T: número, nome/usuário, grupo, valor, WhatsApp."""
    __tablename__ = "phone_lines"
    
    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(30), nullable=False, index=True)
    name = db.Column(db.String(200), index=True)
    group_name = db.Column(db.String(30), index=True)
    amount = db.Column(db.Float, default=0)
    has_whatsapp = db.Column(db.Boolean, default=False)
    notes = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class BankStatementSummary(db.Model):
    """BOA statement summary for Home display."""
    __tablename__ = "bank_statement_summaries"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="CASCADE"), nullable=True, index=True)
    beginning_balance = db.Column(db.Float, default=0)
    total_credits = db.Column(db.Float, default=0)
    total_debits = db.Column(db.Float, default=0)
    ending_balance = db.Column(db.Float, default=0)
    statement_date = db.Column(db.String(80))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    # JSON: credit_groups, debit_groups, transactions (últimas linhas), date_range
    details_json = db.Column(db.Text, nullable=True)
    user = db.relationship("User", backref="bank_statements")


class NewsPost(db.Model):
    """Feed interno estilo Twitter — máx. 60 caracteres no app."""
    __tablename__ = "news_posts"

    id = db.Column(db.Integer, primary_key=True)
    body = db.Column(db.String(60), nullable=False)
    author_name = db.Column(db.String(80), nullable=False, default="Equipe")
    author_role = db.Column(db.String(80), default="")
    category = db.Column(db.String(32), nullable=False, default="Geral")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    link_url = db.Column(db.String(500), nullable=True)
    process_area_tag = db.Column(db.String(40), nullable=True)


class NewsReadReceipt(db.Model):
    __tablename__ = "news_read_receipts"

    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey("news_posts.id", ondelete="CASCADE"), nullable=False, index=True)
    user_key = db.Column(db.String(64), nullable=False)
    reader_name = db.Column(db.String(120), nullable=False)
    read_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)

    __table_args__ = (db.UniqueConstraint("post_id", "user_key", name="uq_news_read_post_user"),)


class NewsReaction(db.Model):
    __tablename__ = "news_reactions"

    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey("news_posts.id", ondelete="CASCADE"), nullable=False, index=True)
    user_key = db.Column(db.String(64), nullable=False)
    emoji = db.Column(db.String(16), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (db.UniqueConstraint("post_id", "user_key", "emoji", name="uq_news_react_user_emoji"),)


class ProcessLibraryItem(db.Model):
    """Biblioteca de processos (PDF) — News & Updates."""
    __tablename__ = "process_library_items"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    area = db.Column(db.String(40), nullable=False, index=True)
    version = db.Column(db.String(24), default="v1.0")
    responsible = db.Column(db.String(120), default="")
    stored_filename = db.Column(db.String(160), nullable=False)
    original_filename = db.Column(db.String(255), default="")
    file_size_bytes = db.Column(db.Integer, default=0)
    description = db.Column(db.String(120), default="")
    uploaded_by_user_id = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    uploaded_by_name = db.Column(db.String(120), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class LongTermRollSnapshot(db.Model):
    """Último rent roll unificado (tenant_directory + rent_roll + auxiliares) — Long Term."""

    __tablename__ = "long_term_roll_snapshots"

    id = db.Column(db.Integer, primary_key=True)
    payload_json = db.Column(db.Text, nullable=False, default="{}")
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class CarsInsuranceSnapshot(db.Model):
    """Último processamento Cars & Insurance (linhas deduplicadas por VIN + data de seguro)."""

    __tablename__ = "cars_insurance_snapshots"

    id = db.Column(db.Integer, primary_key=True)
    payload_json = db.Column(db.Text, nullable=False, default="{}")
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, index=True)


class ModuleSnapshot(db.Model):
    """Snapshot genérico de qualquer módulo — 1 registo por module_key."""
    __tablename__ = "module_snapshots"
    id           = db.Column(db.Integer, primary_key=True)
    module_key   = db.Column(db.String(60), nullable=False, unique=True, index=True)
    payload_json = db.Column(db.Text, nullable=False, default="{}")
    updated_at   = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, index=True)
    saved_by     = db.Column(db.String(120), default="")


class LongTermRollUploadLog(db.Model):
    __tablename__ = "long_term_roll_upload_logs"

    id = db.Column(db.Integer, primary_key=True)
    mode = db.Column(db.String(16), default="replace")
    user_label = db.Column(db.String(120), default="")
    filenames = db.Column(db.String(600), default="")
    rows_imported = db.Column(db.Integer, default=0)
    validation_ok = db.Column(db.Boolean, default=True)
    message = db.Column(db.String(500), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class LongTermRollRowAction(db.Model):
    """Ação por linha (APROVAR / RECUSAR / REVER)."""

    __tablename__ = "long_term_roll_row_actions"

    row_key = db.Column(db.String(64), primary_key=True)
    action = db.Column(db.String(24), nullable=False)
    user_label = db.Column(db.String(120), default="")
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class LongTermRollAuditLog(db.Model):
    __tablename__ = "long_term_roll_audit_logs"

    id = db.Column(db.Integer, primary_key=True)
    row_key = db.Column(db.String(64), index=True, nullable=False)
    action = db.Column(db.String(24), nullable=False)
    user_label = db.Column(db.String(120), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


PROCESS_LIBRARY_AREAS = [
    "Financeiro",
    "Reservas",
    "Limpeza",
    "Manutenção",
    "RH/Payroll",
    "Jurídico",
    "TI",
    "Geral",
]


def _process_area_to_news_category(area):
    return {
        "Financeiro": "Statement",
        "Reservas": "Geral",
        "Limpeza": "Geral",
        "Manutenção": "Geral",
        "RH/Payroll": "Payroll",
        "Jurídico": "Urgente",
        "TI": "Geral",
        "Geral": "Geral",
    }.get(area, "Geral")


def _is_crm_admin():
    uid = session.get("user_id")
    if not uid:
        return False
    u = User.query.get(uid)
    return bool(u and u.is_admin)


def _can_manage_process_item(proc):
    if _is_crm_admin():
        return True
    uid = session.get("user_id")
    if uid and proc.uploaded_by_user_id and int(proc.uploaded_by_user_id) == int(uid):
        return True
    return False


def _serialize_process_library_item(proc):
    now = datetime.utcnow()
    age_sec = (now - proc.created_at).total_seconds() if proc.created_at else 999999
    is_novo = age_sec < 48 * 3600
    return {
        "id": proc.id,
        "name": proc.name,
        "area": proc.area,
        "version": proc.version or "v1.0",
        "responsible": proc.responsible or "—",
        "description": proc.description or "",
        "file_size_kb": round((proc.file_size_bytes or 0) / 1024.0, 1),
        "created_at": proc.created_at.strftime("%Y-%m-%dT%H:%M:%S") + "Z" if proc.created_at else "",
        "uploaded_by_name": proc.uploaded_by_name or "—",
        "view_url": url_for("crm_process_library_file", pid=proc.id),
        "download_url": url_for("crm_process_library_download", pid=proc.id),
        "update_url": url_for("crm_process_library_update", pid=proc.id),
        "delete_url": url_for("crm_process_library_delete", pid=proc.id),
        "is_novo": is_novo,
        "can_edit": _can_manage_process_item(proc),
        "can_delete": _can_manage_process_item(proc),
    }


def _process_library_grouped_models():
    all_procs = ProcessLibraryItem.query.order_by(ProcessLibraryItem.created_at.desc()).all()
    grouped = {a: [] for a in PROCESS_LIBRARY_AREAS}
    for p in all_procs:
        bucket = p.area if p.area in PROCESS_LIBRARY_AREAS else "Geral"
        grouped[bucket].append(p)
    return all_procs, grouped


def _process_library_api_summary_dict():
    all_procs, grouped = _process_library_grouped_models()
    last = all_procs[0] if all_procs else None
    return {
        "last": _serialize_process_library_item(last) if last else None,
        "total": len(all_procs),
        "by_area": {k: [_serialize_process_library_item(p) for p in v] for k, v in grouped.items()},
        "items": [_serialize_process_library_item(p) for p in all_procs],
    }


def _process_library_template_context():
    all_procs, grouped = _process_library_grouped_models()
    last = all_procs[0] if all_procs else None
    pl_now = datetime.utcnow()
    process_item_meta = {}
    for p in all_procs:
        age = (pl_now - p.created_at).total_seconds() if p.created_at else 999999
        process_item_meta[p.id] = {"is_novo": age < 48 * 3600}
    users = User.query.filter_by(is_active=True).order_by(User.username).limit(300).all()
    team_select = []
    for u in users:
        disp = " ".join(x for x in (u.first_name, u.last_name) if x).strip() or u.username or u.email or str(u.id)
        team_select.append({"id": u.id, "name": disp[:120], "initial": (disp[:1] or "?").upper()})
    if not team_select:
        team_select = [
            {"id": 0, "name": "Equipe CFO", "initial": "E"},
            {"id": 0, "name": "Controller", "initial": "C"},
            {"id": 0, "name": "Operações Reservas", "initial": "O"},
        ]
    serialized = [_serialize_process_library_item(p) for p in all_procs]
    process_can_manage = {p.id: _can_manage_process_item(p) for p in all_procs}
    return {
        "process_areas": PROCESS_LIBRARY_AREAS,
        "processes_by_area": grouped,
        "process_total": len(all_procs),
        "last_process": last,
        "process_team_select": team_select,
        "process_library_json": json.dumps(serialized, ensure_ascii=False),
        "last_process_json": json.dumps(_serialize_process_library_item(last), ensure_ascii=False) if last else "null",
        "process_summary_url": url_for("crm_process_library_api_summary"),
        "process_upload_url": url_for("crm_process_library_upload"),
        "pl_now": pl_now,
        "process_item_meta": process_item_meta,
        "process_can_manage": process_can_manage,
    }


class CrmServiceRecord(db.Model):
    """Registros institucionais: PW, Pool, Land Scape (campo module)."""
    __tablename__ = "crm_service_records"

    id = db.Column(db.Integer, primary_key=True)
    module = db.Column(db.String(24), nullable=False, index=True)
    record_date = db.Column(db.Date, index=True)
    property_name = db.Column(db.String(200), index=True, default="")
    category = db.Column(db.String(120), index=True, default="")
    owner = db.Column(db.String(200), index=True, default="")
    status = db.Column(db.String(80), index=True, default="")
    service_type = db.Column(db.String(120), index=True, default="")
    amount = db.Column(db.Float, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class CrmModuleUploadLog(db.Model):
    """Log de uploads CSV por módulo."""
    __tablename__ = "crm_module_upload_logs"

    id = db.Column(db.Integer, primary_key=True)
    module = db.Column(db.String(24), nullable=False, index=True)
    action = db.Column(db.String(20))  # replace, append
    filename = db.Column(db.String(255))
    rows_imported = db.Column(db.Integer, default=0)
    validation_ok = db.Column(db.Boolean, default=True)
    message = db.Column(db.String(500))
    user_label = db.Column(db.String(120))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class CrmModuleViewLog(db.Model):
    """Log de visualização das páginas dos módulos."""
    __tablename__ = "crm_module_view_logs"

    id = db.Column(db.Integer, primary_key=True)
    module = db.Column(db.String(24), nullable=False, index=True)
    user_label = db.Column(db.String(120))
    ip_address = db.Column(db.String(64))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class PoolDashboardSnapshot(db.Model):
    """Último snapshot do dashboard Pool (planilha Excel mensal)."""

    __tablename__ = "pool_dashboard_snapshots"

    id = db.Column(db.Integer, primary_key=True)
    period_key = db.Column(db.String(20), index=True, default="")
    payload_json = db.Column(db.Text)
    confirmation = db.Column(db.String(500), default="")
    filename = db.Column(db.String(255), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class IntegrationConnection(db.Model):
    """Conexão OAuth / API por provider (tokens cifrados com Fernet + SECRET_KEY)."""

    __tablename__ = "integration_connections"

    id = db.Column(db.Integer, primary_key=True)
    provider = db.Column(db.String(80), nullable=False, index=True)
    label = db.Column(db.String(200), default="")
    external_account_id = db.Column(db.String(255), nullable=True, index=True)
    status = db.Column(db.String(20), default="active", index=True)
    oauth_access_token_encrypted = db.Column(db.Text)
    oauth_refresh_token_encrypted = db.Column(db.Text)
    oauth_expires_at = db.Column(db.DateTime)
    extra_config_encrypted = db.Column(db.Text)
    webhook_secret_encrypted = db.Column(db.Text)
    last_sync_at = db.Column(db.DateTime)
    last_error = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    logs = db.relationship("IntegrationLog", backref="connection", lazy="dynamic")

    __table_args__ = (
        db.UniqueConstraint("provider", "external_account_id", name="uq_integration_provider_account"),
    )


class IntegrationLog(db.Model):
    """Audit trail de sync, webhooks e OAuth."""

    __tablename__ = "integration_logs"

    id = db.Column(db.Integer, primary_key=True)
    connection_id = db.Column(db.Integer, db.ForeignKey("integration_connections.id", ondelete="SET NULL"), nullable=True, index=True)
    provider = db.Column(db.String(80), nullable=False, index=True)
    action = db.Column(db.String(80), nullable=False, index=True)
    status = db.Column(db.String(20), default="ok", index=True)
    message = db.Column(db.Text)
    details_json = db.Column(db.Text)
    duration_ms = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class IntegrationExternalRecord(db.Model):
    """Chave idempotente provider + external_id + tipo de registo."""

    __tablename__ = "integration_external_records"

    id = db.Column(db.Integer, primary_key=True)
    provider = db.Column(db.String(80), nullable=False, index=True)
    external_id = db.Column(db.String(255), nullable=False, index=True)
    record_type = db.Column(db.String(80), default="generic", index=True)
    connection_id = db.Column(db.Integer, db.ForeignKey("integration_connections.id", ondelete="CASCADE"), nullable=True, index=True)
    payload_hash = db.Column(db.String(64))
    last_seen_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint("provider", "external_id", "record_type", name="uq_integration_external_triple"),
    )


class IntegrationSyncJob(db.Model):
    """Fila de sync — Celery pode preencher celery_task_id; sem worker fica pending/failed manual."""

    __tablename__ = "integration_sync_jobs"

    id = db.Column(db.Integer, primary_key=True)
    connection_id = db.Column(db.Integer, db.ForeignKey("integration_connections.id", ondelete="CASCADE"), nullable=False, index=True)
    job_type = db.Column(db.String(80), default="full_sync")
    status = db.Column(db.String(20), default="pending", index=True)
    celery_task_id = db.Column(db.String(120))
    error_message = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    started_at = db.Column(db.DateTime)
    finished_at = db.Column(db.DateTime)


class IntegrationWebhookDelivery(db.Model):
    """Últimas entregas de webhook (debug / compliance)."""

    __tablename__ = "integration_webhook_deliveries"

    id = db.Column(db.Integer, primary_key=True)
    provider = db.Column(db.String(80), nullable=False, index=True)
    event_id = db.Column(db.String(120), index=True)
    signature_ok = db.Column(db.Boolean, default=False)
    body_preview = db.Column(db.String(500))
    processed_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class Reserva(db.Model):
    """Reservas - dados de reservas importados via CSV."""
    __tablename__ = "reservas"
    
    id = db.Column(db.Integer, primary_key=True)
    reservation_id = db.Column(db.String(80), index=True)
    source = db.Column(db.String(120), index=True)
    unit = db.Column(db.String(120), index=True)
    guest = db.Column(db.String(200))
    reservation_date = db.Column(db.Date, index=True)
    nights = db.Column(db.Float, default=0)
    total_inc_tax = db.Column(db.Float, default=0)
    amount_received = db.Column(db.Float, default=0)
    property_name = db.Column(db.String(200), index=True)
    owner = db.Column(db.String(200), index=True)
    status = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class RegionsWeUnit(db.Model):
    """Unidade/casa para W & E Regional Overview (agregação por Community)."""
    __tablename__ = "regions_we_units"

    id = db.Column(db.Integer, primary_key=True)
    community = db.Column(db.String(200), nullable=False, index=True)
    region = db.Column(db.String(20), index=True)  # W, E
    bedrooms = db.Column(db.Integer, default=0)
    owner = db.Column(db.String(200), default="")
    unit_name = db.Column(db.String(200), default="")
    is_out = db.Column(db.Boolean, default=False, index=True)  # fora da operação / marcado Out na planilha
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


def _regions_we_find_col(fieldnames, *candidates):
    if not fieldnames:
        return None
    fn = [str(f or "").strip() for f in fieldnames if f is not None and str(f).strip()]
    lower_map = {re.sub(r"\s+", " ", f.lower()): f for f in fn}
    for cand in candidates:
        key = re.sub(r"\s+", " ", str(cand).lower())
        if key in lower_map:
            return lower_map[key]
    for cand in candidates:
        ck = re.sub(r"\s+", "", str(cand).lower())
        for f in fn:
            fk = re.sub(r"\s+", "", f.lower())
            if ck and ck in fk:
                return f
    return None


def _regions_we_parse_region_cell(val):
    s = str(val or "").strip().upper()
    if not s:
        return ""
    if "WE" in s or (("W" in s) and ("E" in s) and len(s) <= 8):
        return "WE"
    if s in ("W", "WEST", "OESTE") or s.startswith("W ") or " WEST" in f" {s}" or "REGION W" in s:
        return "W"
    if s in ("E", "EAST", "LESTE") or s.startswith("E ") or " EAST" in f" {s}" or "REGION E" in s:
        return "E"
    if "WEST" in s:
        return "W"
    if "EAST" in s:
        return "E"
    if "W" in s and "E" not in s:
        return "W"
    if "E" in s:
        return "E"
    return ""


def _regions_we_rows_from_dicts(fieldnames, dict_rows):
    col_comm = _regions_we_find_col(
        fieldnames, "community", "comunidade", "condominio", "condomínio", "community name", "base"
    )
    col_reg = _regions_we_find_col(fieldnames, "regiao", "região", "region", "reg", "area", "zona")
    col_bed = _regions_we_find_col(fieldnames, "bedrooms", "bed", "quartos", "bedroom", "beds", "quarto")
    col_owner = _regions_we_find_col(fieldnames, "owner", "proprietario", "proprietário", "landlord")
    col_unit = _regions_we_find_col(
        fieldnames, "unit name", "unit", "unidade", "property", "casa", "home", "nome da unidade"
    )
    col_out = _regions_we_find_col(
        fieldnames,
        "out",
        "outs",
        "is_out",
        "fora",
        "status out",
        "situacao",
        "situação",
        "fuera",
    )
    if not col_comm:
        return [], ["Não encontrei coluna Community / Condomínio no CSV. Inclua um cabeçalho reconhecível."]
    out = []
    for raw in dict_rows:
        if not raw:
            continue
        row = {str(k).strip(): v for k, v in raw.items() if k is not None and str(k).strip()}
        comm = str(row.get(col_comm) or "").strip()
        if not comm:
            continue
        reg = _regions_we_parse_region_cell(row.get(col_reg) if col_reg else "")
        if not reg:
            reg = "W"
        bed = 0
        if col_bed and row.get(col_bed) not in (None, ""):
            try:
                bed = int(float(str(row.get(col_bed)).replace(",", ".").strip()))
            except (TypeError, ValueError):
                bed = 0
        owner = str(row.get(col_owner) or "").strip() if col_owner else ""
        unit = str(row.get(col_unit) or "").strip() if col_unit else ""
        is_out_row = False
        if col_out:
            ov = row.get(col_out)
            if ov is not None and ov != "":
                vs = str(ov).strip().upper()
                if isinstance(ov, (int, float)) and int(ov) == 1:
                    is_out_row = True
                elif vs in (
                    "Y",
                    "YES",
                    "S",
                    "SIM",
                    "TRUE",
                    "1",
                    "OUT",
                    "OUTS",
                    "X",
                    "*",
                    "FORA",
                    "FUERA",
                ):
                    is_out_row = True
        out.append(
            {
                "community": comm[:200],
                "region": reg[:20],
                "bedrooms": max(0, bed),
                "owner": owner[:200],
                "unit_name": unit[:200],
                "is_out": is_out_row,
            }
        )
    if not out:
        return [], ["Nenhuma linha com Community preenchida."]
    return out, []


def _fmt_currency(val):
    """Format value as currency for templates."""
    if val is None:
        return "$ 0,00"
    try:
        d = float(val)
        s = f"{d:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"$ {s}"
    except (TypeError, ValueError):
        return "$ 0,00"


def _fmt_hours(val):
    """Format hours (contractors / payroll)."""
    if val is None:
        return "0.00"
    try:
        return f"{float(val):.2f}"
    except (TypeError, ValueError):
        return "0.00"


def _crm_distinct_companies_count(module_key: str) -> int:
    """Empresas nos módulos PW / Pool / Land Scape = valores distintos em category."""
    rows = (
        db.session.query(CrmServiceRecord.category)
        .filter(CrmServiceRecord.module == module_key)
        .distinct()
        .all()
    )
    return len({str(c[0]).strip() for c in rows if c[0] and str(c[0]).strip() not in ("", "—")})


def _contractors_unique_company_count(csv_rows) -> int:
    if not csv_rows:
        return 0
    seen = set()
    for r in csv_rows:
        if isinstance(r, dict):
            name = r.get("contractor_name")
        else:
            name = getattr(r, "contractor_name", None)
        n = (str(name or "").strip().lower())
        if n:
            seen.add(n)
    return len(seen)


def _payroll_distinct_department_count(payroll_rows, employees) -> int:
    """Para Payroll, 'empresas' = departamentos distintos com dados."""
    seen = set()
    for row in payroll_rows or []:
        emp = getattr(row, "employee", None) or (row.get("employee") if isinstance(row, dict) else None)
        if emp is None:
            continue
        d = getattr(emp, "department", None) or (emp.get("department") if isinstance(emp, dict) else None)
        ds = str(d or "").strip().lower()
        if ds and ds != "—":
            seen.add(ds)
    for e in employees or []:
        d = getattr(e, "department", None) or (e.get("department") if isinstance(e, dict) else None)
        ds = str(d or "").strip().lower()
        if ds and ds != "—":
            seen.add(ds)
    return len(seen)


def _display_statement_from_summary(summary):
    """Monta dict compatível com statement.html a partir do modelo + JSON gravado no upload."""
    if not summary:
        return None
    data = {
        "beginning_balance": float(summary.beginning_balance or 0),
        "total_credits": float(summary.total_credits or 0),
        "total_debits": float(summary.total_debits or 0),
        "ending_balance": float(summary.ending_balance or 0),
        "balance": float(summary.ending_balance or 0),
        "date_range": summary.statement_date or "",
        "credit_groups": {},
        "debit_groups": {},
        "transactions": [],
    }
    if summary.details_json:
        try:
            extra = json.loads(summary.details_json)
            data["credit_groups"] = extra.get("credit_groups") or {}
            data["debit_groups"] = extra.get("debit_groups") or {}
            data["transactions"] = extra.get("transactions") or []
            if extra.get("date_range"):
                data["date_range"] = extra["date_range"]
        except (json.JSONDecodeError, TypeError):
            pass
    return data


def _news_session_user_key():
    if not session.get("news_reader_key"):
        session["news_reader_key"] = uuid.uuid4().hex
    return session["news_reader_key"]


def _news_reader_display_name():
    u = current_user()
    if u:
        parts = [u.first_name or "", u.last_name or ""]
        name = " ".join(p for p in parts if p).strip()
        return (name or u.username or u.email or "User")[:120]
    return "Visitante"


def _integration_fernet(secret_key: str):
    """Fernet derivado de SECRET_KEY — tokens OAuth em repouso cifrados."""
    from cryptography.fernet import Fernet

    key = base64.urlsafe_b64encode(hashlib.sha256(secret_key.encode("utf-8")).digest())
    return Fernet(key)


def integration_encrypt_token(secret_key: str, plain: str) -> str:
    if not plain:
        return ""
    return _integration_fernet(secret_key).encrypt(plain.encode("utf-8")).decode("ascii")


def integration_decrypt_token(secret_key: str, blob: str) -> str:
    if not blob:
        return ""
    return _integration_fernet(secret_key).decrypt(blob.encode("ascii")).decode("utf-8")


def integration_verify_webhook_signature(provider: str, raw_body: bytes, headers: dict) -> bool:
    """
    Verificação de assinatura por provider. Sem segredo configurado → True (dev).
    Produção: definir env por provider (ex. RAMP_WEBHOOK_SECRET).
    """
    p = (provider or "").lower().strip()
    hdr = {k.lower(): v for k, v in (headers or {}).items()}
    if p == "quickbooks" or p == "qbo":
        token = os.getenv("QBO_WEBHOOK_VERIFIER_TOKEN", "").strip()
        if not token:
            return True
        sig = hdr.get("intuit-signature") or ""
        try:
            digest = hmac.new(token.encode("utf-8"), raw_body, hashlib.sha256).digest()
            expected = base64.b64encode(digest).decode("ascii")
            return hmac.compare_digest(sig.strip(), expected)
        except Exception:
            return False
    env_key = f"{p.upper()}_WEBHOOK_SECRET"
    secret = os.getenv(env_key, "").strip()
    if not secret:
        return True
    sig_hdr = hdr.get("x-signature") or hdr.get("x-hub-signature-256") or hdr.get("x-ramp-signature") or ""
    if "sha256=" in sig_hdr:
        sig_hdr = sig_hdr.split("sha256=", 1)[-1].strip()
    try:
        digest = hmac.new(secret.encode("utf-8"), raw_body, hashlib.sha256).hexdigest()
        return hmac.compare_digest(digest, sig_hdr) or hmac.compare_digest(
            base64.b64encode(hmac.new(secret.encode("utf-8"), raw_body, hashlib.sha256).digest()).decode("ascii"),
            sig_hdr,
        )
    except Exception:
        return False


INTEGRATION_PROVIDERS = (
    ("ramp", "Ramp", "Corporate cards & spend"),
    ("rentengine", "RentEngine", "Leasing / listings pipeline"),
    ("quickbooks", "QuickBooks", "Accounting & bills (Intuit)"),
    ("mls", "MLS", "Listings feed"),
    ("billcom", "Bill.com", "AP / payments"),
    ("boom", "Boom", "Workflow / automation"),
    ("appfolio", "AppFolio", "Property management"),
)


def create_app():
    """Application factory."""
    # Pasta deste ficheiro (website/): templates, static e instance ficam ao lado de app.py
    root_dir = os.path.dirname(os.path.abspath(__file__))
    app = Flask(__name__,
                template_folder=os.path.join(root_dir, "templates"),
                static_folder=os.path.join(root_dir, "static"))
    app.jinja_env.filters["fmt_currency"] = _fmt_currency
    app.jinja_env.filters["fmt_hours"] = _fmt_hours

    # .env na pasta website/ e na raiz do GodManager (QuickBooks, etc.)
    try:
        from dotenv import load_dotenv

        load_dotenv(os.path.join(root_dir, ".env"))
        load_dotenv(os.path.join(os.path.dirname(root_dir), ".env"))
    except Exception:
        pass

    # Configuration
    app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "your-secret-key-change-in-production")
    app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv(
        "DATABASE_URL",
        "sqlite:///master_finance.db"  # SQLite for local dev; use PostgreSQL in production
    )
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    _proc_store = os.path.join(root_dir, "instance", "process_library")
    os.makedirs(_proc_store, exist_ok=True)
    app.config["PROCESS_LIBRARY_STORAGE"] = _proc_store
    
    # Redis session configuration
    redis_url = os.getenv("REDIS_URL", "redis://:redispassword@redis:6379/0")
    app.config["SESSION_TYPE"] = "redis"
    app.config["SESSION_REDIS"] = redis_url
    app.config["SESSION_COOKIE_SECURE"] = False  # True bloqueia cookie em localhost (HTTP)
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
    app.config["SESSION_PERMANENT"] = False
    
    # Parse Redis URL if needed for Flask-Session
    try:
        from redis import Redis
        parsed = urlparse(redis_url)
        password = parsed.password if parsed.password else None
        app.config["SESSION_REDIS"] = Redis(
            host=parsed.hostname or "redis",
            port=parsed.port or 6379,
            password=password,
            db=int(parsed.path.strip("/")) if parsed.path else 0,
            decode_responses=False
        )
    except Exception as e:
        # Fallback to filesystem sessions if Redis unavailable
        app.config["SESSION_TYPE"] = "filesystem"
        app.config["SESSION_REDIS"] = None
    
    # Initialize extensions
    db.init_app(app)
    sess.init_app(app)
    
    # Register routes
    register_routes(app)
    
    # Create tables
    with app.app_context():
        db.create_all()
        # Migration: add new columns to phone_lines if missing
        for col, typ in [("name", "VARCHAR(200)"), ("group_name", "VARCHAR(30)"), ("amount", "FLOAT")]:
            try:
                db.session.execute(text(f"ALTER TABLE phone_lines ADD COLUMN {col} {typ}"))
                db.session.commit()
            except Exception:
                db.session.rollback()
        # Migration: bank_statement_summaries.details_json (SQLite não atualiza schema no create_all)
        try:
            db.session.execute(
                text("ALTER TABLE bank_statement_summaries ADD COLUMN details_json TEXT")
            )
            db.session.commit()
        except Exception:
            db.session.rollback()
        for stmt in (
            "ALTER TABLE news_posts ADD COLUMN link_url VARCHAR(500)",
            "ALTER TABLE news_posts ADD COLUMN process_area_tag VARCHAR(40)",
        ):
            try:
                db.session.execute(text(stmt))
                db.session.commit()
            except Exception:
                db.session.rollback()
        try:
            db.session.execute(text("ALTER TABLE regions_we_units ADD COLUMN is_out INTEGER DEFAULT 0"))
            db.session.commit()
        except Exception:
            db.session.rollback()
    
    return app


def _normalize_source(s):
    if not s: return ""
    return re.sub(r'\s+', ' ', str(s).strip())

def _parse_agencia_date(val):
    if not val: return None
    s = str(val).strip()
    if not s: return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None

def _parse_agencia_number(val):
    if val is None or val == '': return Decimal('0')
    s = str(val).strip().replace(',', '').replace(' ', '')
    s = re.sub(r'[^\d.\-]', '', s)
    try: return Decimal(s) if s else Decimal('0')
    except: return Decimal('0')

def _format_agencia_currency(val):
    if val is None: return "$ 0,00"
    d = float(val)
    s = f"{d:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"$ {s}"


def _crm_cars_insurance_row_ins_date(row):
    """Extrai date a partir de insDate ISO (YYYY-MM-DD) no payload salvo."""
    if not isinstance(row, dict):
        return None
    raw = row.get("insDate")
    if not raw or not isinstance(raw, str) or len(raw) < 10:
        return None
    try:
        y, m, d = int(raw[:4]), int(raw[5:7]), int(raw[8:10])
        return date(y, m, d)
    except (ValueError, TypeError):
        return None


def _crm_cars_insurance_home_stats():
    """
    KPIs para Home a partir do último snapshot Cars & Insurance.
    Status alinhado ao dashboard (date local do servidor).
    """
    snap = CarsInsuranceSnapshot.query.order_by(CarsInsuranceSnapshot.updated_at.desc()).first()
    if not snap or not (snap.payload_json or "").strip():
        return None
    try:
        data = json.loads(snap.payload_json)
    except (json.JSONDecodeError, TypeError):
        return None
    rows = data.get("rows")
    if not isinstance(rows, list):
        return None
    today = date.today()
    t30 = today + timedelta(days=30)
    total = 0
    exp = valido = exp30 = sem = 0
    ativos = 0
    for r in rows:
        if not isinstance(r, dict):
            continue
        total += 1
        ins_d = _crm_cars_insurance_row_ins_date(r)
        if ins_d is None:
            sem += 1
            ativos += 1
        elif ins_d < today:
            exp += 1
        elif ins_d <= t30:
            exp30 += 1
            ativos += 1
        else:
            valido += 1
            ativos += 1
    # Seguros com data de vigência/vencimento (linhas com insDate); o resto conta só como frota
    seguros = total - sem
    return {
        "total": total,
        "seguros": seguros,
        "ativos": ativos,
        "expirados": exp,
        "valido": valido,
        "exp30": exp30,
        "sem_data": sem,
        "updated_at": snap.updated_at,
    }


def get_client_ip():
    """Get client IP address."""
    if request.headers.get("X-Forwarded-For"):
        return request.headers.get("X-Forwarded-For").split(",")[0].strip()
    return request.remote_addr or "unknown"


def get_user_agent():
    """Get user agent string."""
    return request.headers.get("User-Agent", "unknown")


def current_user():
    """Get current logged-in user."""
    user_id = session.get("user_id")
    if not user_id:
        return None
    return User.query.get(user_id)


def _lt_actor_label():
    u = current_user()
    if u:
        return (u.username or str(u.id))[:120]
    if session.get("news_reader_key"):
        return session["news_reader_key"][:120]
    return (get_client_ip() or "visitante")[:120]


def _register_crm_routes(app):
    """Register CRM routes for Master Finance dashboard."""

    @app.route("/crm")
    @app.route("/crm/")
    def crm_home():
        from datetime import datetime
        from website.translations import t as _crm_t

        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
        # KPIs para todos os módulos do menu (quadrado + donut)
        _url = url_for
        home_stat_charts = [
            {"label": "Properties", "value": "0", "prefix": "prop", "labels_json": "[\"W\", \"E\"]", "values_json": "[0, 0]", "link": _url("crm_propertys"), "link_text": "Ver Properties", "hide_mini_chart": True, "properties_dashboard": True},
            {"label": "Reservas", "value": "0", "prefix": "reservas", "labels_json": "[\"N/A\"]", "values_json": "[0]", "link": _url("crm_reservas"), "link_text": "Ver Reservas"},
            {"label": "Renovations", "value": "$0", "prefix": "reno", "labels_json": "[\"—\"]", "values_json": "[0]", "link": _url("crm_renovations"), "link_text": "Ver Renovations", "hide_mini_chart": True, "renovations_dashboard": True},
            {"label": "Cleaners", "value": "0", "prefix": "clean", "labels_json": "[\"Limpezas\", \"Valor\"]", "values_json": "[0, 0]", "link": _url("crm_cleaners"), "link_text": "Ver Cleaners", "companies_count": 0, "hk_sync": True},
            {"label": "Total Cleaners Value", "value": "$ 0,00", "prefix": "hkval", "labels_json": "[\"Total\"]", "values_json": "[0]", "link": _url("crm_cleaners"), "link_text": "Ver Cleaners", "hide_mini_chart": True, "hk_value_card": True},
            {"label": "Contractors", "value": "0", "prefix": "contr", "labels_json": "[\"Ativos\"]", "values_json": "[0]", "link": _url("crm_contractors"), "link_text": "Ver Contractors"},
            {"label": "Payroll", "value": "$ 0,00", "prefix": "pay", "labels_json": "[\"N/A\"]", "values_json": "[0]", "link": _url("crm_payroll"), "link_text": "Ver Payroll"},
            {"label": "Licenses", "value": "0", "prefix": "lic", "labels_json": "[\"Em dia\", \"A vencer\"]", "values_json": "[0, 0]", "link": _url("crm_licenses"), "link_text": "Ver Licenses"},
            {"label": "Stock", "value": "0", "prefix": "stock", "labels_json": "[\"Itens\"]", "values_json": "[0]", "link": _url("crm_stock"), "link_text": "Ver Stock"},
            {"label": "Cartões", "value": "0", "prefix": "cart", "labels_json": "[\"Ativos\"]", "values_json": "[0]", "link": _url("crm_cartoes"), "link_text": "Ver Cartões"},
            {"label": "Cars", "value": "0", "prefix": "cars", "labels_json": "[\"Ativos\"]", "values_json": "[0]", "link": _url("crm_cars"), "link_text": "Ver Cars"},
            {"label": "Insurance", "value": "0", "prefix": "ins", "labels_json": "[\"Políticas\"]", "values_json": "[0]", "link": _url("crm_insurance"), "link_text": "Ver Insurance"},
            {"label": "Long Term", "value": "0", "prefix": "lt", "labels_json": "[\"N/A\"]", "values_json": "[0]", "link": _url("crm_long_term"), "link_text": "Ver Long Term"},
            {"label": "Agências", "value": "0", "prefix": "agen", "labels_json": "[\"Total\"]", "values_json": "[0]", "link": _url("crm_agencias"), "link_text": "Ver Agências"},
            {"label": "Vendor", "value": "0", "prefix": "vend", "labels_json": "[\"Ativos\", \"Inativos\"]", "values_json": "[0, 0]", "link": _url("crm_vendor"), "link_text": "Ver Vendor"},
            {"label": "Cofre", "value": "$ 0,00", "prefix": "cofre", "is_cofre": True, "link": _url("crm_cofre")},
            {"label": "Upcoming", "value": "$0", "prefix": "upc", "labels_json": "[\"Pendentes\"]", "values_json": "[0]", "link": _url("crm_upcoming"), "link_text": "Ver Upcoming", "hide_mini_chart": True},
            {"label": "1099", "value": "$0.00", "prefix": "1099", "labels_json": "[\"Registros\"]", "values_json": "[0]", "link": _url("crm_1099"), "link_text": "Abrir Report 1099", "hide_mini_chart": True, "irs1099_dashboard": True},
            {"label": "News", "value": "0", "prefix": "news", "labels_json": "[\"Updates\"]", "values_json": "[0]", "link": _url("crm_news"), "link_text": "Ver News"},
            {"label": "Linhas Telefônicas", "value": "0", "prefix": "linhas", "labels_json": "[\"Com WhatsApp\", \"Sem WhatsApp\"]", "values_json": "[0, 0]", "link": _url("crm_linhas_telefonicas"), "link_text": "Ver Linhas"},
            {"label": "PW", "value": "0", "prefix": "pwmod", "labels_json": "[\"N/A\"]", "values_json": "[0]", "link": _url("crm_pw"), "link_text": "Ver PW"},
            {"label": "Pool", "value": "0", "prefix": "poolmod", "labels_json": "[\"N/A\"]", "values_json": "[0]", "link": _url("crm_pool"), "link_text": "Ver Pool"},
            {"label": "Land Scape", "value": "0", "prefix": "lsmod", "labels_json": "[\"N/A\"]", "values_json": "[0]", "link": _url("crm_landscape"), "link_text": "Ver Land Scape"},
        ]
        lt_payload_cache = None
        try:
            _lt_snap = LongTermRollSnapshot.query.order_by(LongTermRollSnapshot.updated_at.desc()).first()
            if _lt_snap and _lt_snap.payload_json and str(_lt_snap.payload_json).strip() not in ("", "{}"):
                lt_payload_cache = json.loads(_lt_snap.payload_json)
        except Exception:
            lt_payload_cache = None
        # Properties: total de casas = módulo Properties (RegionsWeUnit / overview W–E); donut W/WE vs E/WE
        try:
            region_total_w = RegionsWeUnit.query.filter(RegionsWeUnit.region.in_(["W", "WE"])).count()
            region_total_e = RegionsWeUnit.query.filter(RegionsWeUnit.region.in_(["E", "WE"])).count()
            total_casas_home = RegionsWeUnit.query.count()
            total_outs_home = RegionsWeUnit.query.filter_by(is_out=True).count()
            for s in home_stat_charts:
                if s.get("label") == "Properties":
                    s["value"] = str(total_casas_home)
                    sub = _crm_t("properties_home_subtitle")
                    if total_outs_home:
                        sub = sub + " · " + _crm_t("houses_marked_out").format(n=total_outs_home)
                    s["reservas_subtitle"] = sub
                    s["labels_json"] = '["W / WE", "E / WE"]'
                    s["values_json"] = f"[{region_total_w}, {region_total_e}]"
                    break
        except Exception:
            pass
        # Atualiza KPI Linhas Telefônicas com dados reais
        linhas_total = PhoneLine.query.count()
        linhas_wa = PhoneLine.query.filter_by(has_whatsapp=True).count()
        linhas_sem_wa = linhas_total - linhas_wa
        for s in home_stat_charts:
            if s.get("label") == "Linhas Telefônicas":
                s["value"] = str(linhas_total)
                s["labels_json"] = json.dumps([_crm_t("wa_yes"), _crm_t("wa_no")], ensure_ascii=False)
                s["values_json"] = f"[{linhas_wa}, {linhas_sem_wa}]"
                break
        for mk, lbl, pfx in (
            ("pw", "PW", "pwmod"),
            ("pool", "Pool", "poolmod"),
            ("landscape", "Land Scape", "lsmod"),
        ):
            recs = CrmServiceRecord.query.filter_by(module=mk).all()
            n = len(recs)
            c = Counter((r.service_type or "—")[:32] for r in recs)
            top = c.most_common(6)
            if not top:
                lj = json.dumps([_crm_t("no_data")], ensure_ascii=False)
                vj = "[1]"
            else:
                lj = json.dumps([t[0] for t in top], ensure_ascii=False)
                vj = json.dumps([t[1] for t in top])
            n_comp = _crm_distinct_companies_count(mk)
            for s in home_stat_charts:
                if s.get("label") == lbl:
                    s["value"] = str(n)
                    s["prefix"] = pfx
                    s["labels_json"] = lj
                    s["values_json"] = vj
                    s["companies_count"] = n_comp
                    break
        for s in home_stat_charts:
            if s.get("label") == "Vendor":
                s["companies_count"] = 0
            elif s.get("label") == "Contractors":
                s["companies_count"] = 0
            elif s.get("label") == "Payroll":
                s["companies_count"] = 0
        try:
            nposts = NewsPost.query.count()
            nprocs = ProcessLibraryItem.query.count()
            for s in home_stat_charts:
                if s.get("label") == "News":
                    s["value"] = str(nposts)
                    s["labels_json"] = '["Posts", "Processos"]' if nprocs else '["Posts"]'
                    s["values_json"] = f"[{nposts}, {nprocs}]" if nprocs else f"[{nposts}]"
                    if nprocs:
                        s["reservas_subtitle"] = _crm_t("process_library_pdf").format(n=nprocs)
                    break
        except Exception:
            pass
        # Reservas: cartão = quantidade no mês civil atual (UTC); donut = 12 meses do ano civil atual
        res_rows = Reserva.query.all()
        ch_map = {"Airbnb": 0, "Booking": 0, "Marriott": 0, "VRBO": 0, "Outros": 0}
        for r in res_rows:
            src = (r.source or "").upper()
            if "AIRBNB" in src:
                ch_map["Airbnb"] += 1
            elif "BOOKING" in src:
                ch_map["Booking"] += 1
            elif "MARRIOTT" in src:
                ch_map["Marriott"] += 1
            elif "VRBO" in src or "HOMEAWAY" in src:
                ch_map["VRBO"] += 1
            else:
                ch_map["Outros"] += 1
        cy_home = datetime.utcnow().year
        cm_home = datetime.utcnow().month
        mon_labels_home = [_crm_t(f"month_{m}") for m in range(1, 13)]
        month_counts_home = []
        for m in range(1, 13):
            n_m = Reserva.query.filter(
                db.extract("year", Reserva.reservation_date) == cy_home,
                db.extract("month", Reserva.reservation_date) == m,
            ).count()
            month_counts_home.append(n_m)
        res_current_month_home = month_counts_home[cm_home - 1] if month_counts_home else 0
        for s in home_stat_charts:
            if s.get("label") == "Reservas":
                s["value"] = str(res_current_month_home)
                s["reservas_subtitle"] = _crm_t("reservas_home_month_focus").format(
                    n=res_current_month_home,
                    month=_crm_t(f"month_{cm_home}"),
                    year=cy_home,
                )
                s["labels_json"] = json.dumps(mon_labels_home, ensure_ascii=False)
                s["values_json"] = json.dumps(month_counts_home)
                s["link"] = _url("crm_reservas", year=str(cy_home), month=str(cm_home))
                break
        # Long Term: KPIs do último upload (reutiliza payload já carregado)
        try:
            if isinstance(lt_payload_cache, dict) and lt_payload_cache:
                lt_kpi = lt_payload_cache.get("kpis") or {}
                tc = int(lt_kpi.get("total_units_sum") or lt_kpi.get("total_casas") or 0)
                alug = int(lt_kpi.get("casas_alugadas") or 0)
                vag = int(lt_kpi.get("casas_vagas") or 0)
                for s in home_stat_charts:
                    if s.get("label") == "Long Term":
                        s["value"] = str(tc)
                        if alug == 0 and vag == 0:
                            if tc > 0:
                                s["labels_json"] = '["Propriedades"]'
                                s["values_json"] = f"[{tc}]"
                            else:
                                s["labels_json"] = '["N/A"]'
                                s["values_json"] = "[0]"
                        else:
                            s["labels_json"] = json.dumps(["Alugadas", "Vagas"], ensure_ascii=False)
                            s["values_json"] = json.dumps([alug, vag])
                        s["reservas_subtitle"] = _crm_t("units_property_dir_hint")
                        break
        except Exception:
            pass
        # Cars & Insurance: totais do último snapshot gravado no servidor
        try:
            _ci_stats = _crm_cars_insurance_home_stats()
            if _ci_stats and _ci_stats.get("total", 0) > 0:
                _ua = ""
                if _ci_stats.get("updated_at"):
                    _ua = _ci_stats["updated_at"].strftime("%Y-%m-%d %H:%M")
                _lbl_st = json.dumps(
                    ["Válido", "Expira em 30", "Expirado", "Sem data"],
                    ensure_ascii=False,
                )
                _val_st = json.dumps(
                    [
                        _ci_stats["valido"],
                        _ci_stats["exp30"],
                        _ci_stats["expirados"],
                        _ci_stats["sem_data"],
                    ],
                    ensure_ascii=False,
                )
                _sub_saved = _crm_t("cars_insurance_home_saved").format(ua=_ua) if _ua else ""
                _sub_ins = _crm_t("cars_insurance_home_ins_sub").format(
                    total=_ci_stats["total"],
                    ativos=_ci_stats["ativos"],
                    exp=_ci_stats["expirados"],
                )
                for s in home_stat_charts:
                    if s.get("label") == "Cars":
                        s["value"] = str(_ci_stats["total"])
                        s["labels_json"] = _lbl_st
                        s["values_json"] = _val_st
                        if _sub_saved:
                            s["reservas_subtitle"] = _sub_saved
                        break
                for s in home_stat_charts:
                    if s.get("label") == "Insurance":
                        s["value"] = str(_ci_stats["total"])
                        s["labels_json"] = json.dumps(
                            [_crm_t("cars_ins_donut_active"), _crm_t("cars_ins_donut_exp")],
                            ensure_ascii=False,
                        )
                        s["values_json"] = json.dumps(
                            [_ci_stats["ativos"], _ci_stats["expirados"]],
                            ensure_ascii=False,
                        )
                        s["reservas_subtitle"] = _sub_ins
                        break
        except Exception:
            pass
        home_chart_labels = [_crm_t("no_statement_chart")]
        home_chart_values = [1]
        home_chart_total = "—"
        home_chart_donut_title = _crm_t("stmt_chart_sub")
        home_chart_bar_title = _crm_t("reservas_bar_sub")
        stmt_h = BankStatementSummary.query.order_by(BankStatementSummary.created_at.desc()).first()
        if stmt_h:
            tc = float(stmt_h.total_credits or 0)
            td = abs(float(stmt_h.total_debits or 0))
            home_chart_total = _fmt_currency(stmt_h.ending_balance)
            if tc + td > 0:
                home_chart_labels = [_crm_t("credits"), _crm_t("debits")]
                home_chart_values = [tc, td]
        home_bar_labels = list(ch_map.keys())
        home_bar_values = list(ch_map.values())
        _chart_i18n = {
            "Properties": "properties",
            "Reservas": "reservas",
            "Renovations": "renovations",
            "Cleaners": "cleaners",
            "Total Cleaners Value": "cleaners_value",
            "Contractors": "contractors",
            "Payroll": "payroll",
            "Licenses": "licenses",
            "Stock": "stock",
            "Cartões": "cartoes",
            "Cars": "cars",
            "Insurance": "insurance",
            "Long Term": "long_term",
            "Agências": "agencias",
            "Vendor": "vendor",
            "Upcoming": "upcoming_payouts",
            "1099": "tax_1099",
            "News": "news",
            "Linhas Telefônicas": "linha_telefonica",
            "PW": "pw",
            "Pool": "pool",
            "Land Scape": "landscape",
        }
        for s in home_stat_charts:
            lk = _chart_i18n.get(s.get("label"))
            if lk:
                s["display_label"] = _crm_t(lk)
                if s.get("link") and not s.get("is_cofre"):
                    s["link_text"] = _crm_t("view_module") + " " + _crm_t(lk)
        try:
            home_cars_ins = _crm_cars_insurance_home_stats()
        except Exception:
            home_cars_ins = None
        home_cars_snap = None
        if home_cars_ins:
            home_cars_snap = {k: v for k, v in home_cars_ins.items() if k != "updated_at"}
            _ua = home_cars_ins.get("updated_at")
            if _ua is not None:
                try:
                    home_cars_snap["updated_at"] = _ua.strftime("%Y-%m-%d %H:%M")
                except Exception:
                    home_cars_snap["updated_at"] = str(_ua)
        home_prop_seed = {"units": 0, "sub": ""}
        for s in home_stat_charts:
            if s.get("label") == "Properties":
                try:
                    home_prop_seed["units"] = int(s.get("value") or 0)
                except (ValueError, TypeError):
                    home_prop_seed["units"] = 0
                home_prop_seed["sub"] = s.get("reservas_subtitle") or ""
                break
        home_res_seed = {
            "month_count": res_current_month_home,
            "cy": cy_home,
            "cm": cm_home,
            "month_labels": mon_labels_home,
            "month_values": month_counts_home,
        }
        return render_template(
            "crm/home.html",
            home_stat_charts=home_stat_charts,
            home_cars_ins=home_cars_snap,
            home_prop_seed=home_prop_seed,
            home_res_seed=home_res_seed,
            last_updated=now,
            home_chart_labels=home_chart_labels,
            home_chart_values=home_chart_values,
            home_chart_total=home_chart_total,
            home_chart_donut_title=home_chart_donut_title,
            home_chart_bar_title=home_chart_bar_title,
            home_bar_labels=home_bar_labels,
            home_bar_values=home_bar_values,
        )

    def _crm_stub(module, template, **kwargs):
        from datetime import datetime
        return render_template(template, last_updated=datetime.utcnow().strftime("%Y-%m-%d %H:%M"), **kwargs)

    @app.route("/crm/gaap")
    def crm_gaap():
        return render_template("crm/gaap.html")

    ALLOWED_SNAP_MODULES = {
        'reservations','agencies','housekeeper','phone-lines',
        '1099','1099-allbills','contractors','upcoming',
        'properties','payroll-pr','vendor-as',
        'renovations','longterm',
        'cars-home'
    }

    @app.route("/crm/api/snapshot/<module_key>", methods=["GET","POST"])
    def crm_api_snapshot(module_key):
        """GET: devolve snapshot. POST: grava snapshot."""
        if module_key not in ALLOWED_SNAP_MODULES:
            return jsonify({"ok": False, "error": "módulo inválido"}), 400
        if request.method == "GET":
            try:
                s = ModuleSnapshot.query.filter_by(module_key=module_key).first()
                if not s:
                    return jsonify({"ok": True, "data": {}, "updated_at": None, "saved_by": ""})
                data = json.loads(s.payload_json or "{}")
                ua = s.updated_at.isoformat() + "Z" if s.updated_at else None
                return jsonify({"ok": True, "data": data, "updated_at": ua, "saved_by": s.saved_by or ""})
            except Exception as e:
                return jsonify({"ok": False, "error": str(e)}), 500
        # POST
        body = request.get_json(silent=True) or {}
        uid  = session.get("user_id")
        u    = User.query.get(uid) if uid else None
        label = u.username if u else "system"
        try:
            payload = json.dumps(body, ensure_ascii=False, default=str)
            if len(payload.encode()) > 15_000_000:
                return jsonify({"ok": False, "error": "Payload muito grande"}), 413
            s = ModuleSnapshot.query.filter_by(module_key=module_key).first()
            if s:
                s.payload_json = payload
                s.updated_at   = datetime.utcnow()
                s.saved_by     = label
            else:
                s = ModuleSnapshot(module_key=module_key, payload_json=payload, saved_by=label)
                db.session.add(s)
            db.session.commit()
            return jsonify({"ok": True, "saved": module_key, "by": label})
        except Exception as e:
            db.session.rollback()
            return jsonify({"ok": False, "error": str(e)}), 500


    # --- [rota desativada — menu removido] ---
    # @app.route("/crm/statement", methods=["GET", "POST"])
    # def crm_statement():
        # from boa_parser import parse_boa_csv
    #
        # def _decode_csv(raw):
            # for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
                # try:
                    # return raw.decode(enc)
                # except UnicodeDecodeError:
                    # continue
            # return raw.decode("utf-8", errors="replace")
    #
        # summary = BankStatementSummary.query.order_by(BankStatementSummary.created_at.desc()).first()
        # statement_data = None
        # if request.method == "POST" and "csv_file" in request.files:
            # f = request.files["csv_file"]
            # if not f or not f.filename:
                # flash("Selecione um arquivo CSV antes de enviar.", "warning")
            # elif not f.filename.lower().endswith(".csv"):
                # flash("Use um arquivo com extensão .csv (export do Bank of America).", "warning")
            # else:
                # try:
                    # raw = f.read()
                    # content = _decode_csv(raw)
                    # statement_data = parse_boa_csv(content)
                    # warnings = statement_data.get("parse_warnings") or []
                    # ntx = len(statement_data.get("transactions") or [])
                    # if ntx == 0:
                        # flash_msg = "Não foi possível ler transações neste CSV. " + (
                            # warnings[0]
                            # if warnings
                            # else "Verifique se é o extrato detalhado com colunas Date, Description, Amount."
                        # )
                        # flash(flash_msg, "warning")
                    # else:
                        # BankStatementSummary.query.delete()
                        # payload = {
                            # "credit_groups": statement_data.get("credit_groups") or {},
                            # "debit_groups": statement_data.get("debit_groups") or {},
                            # "transactions": (statement_data.get("transactions") or [])[-400:],
                            # "date_range": statement_data.get("date_range") or "",
                        # }
                        # s = BankStatementSummary(
                            # beginning_balance=float(statement_data.get("beginning_balance", 0)),
                            # total_credits=float(statement_data.get("total_credits", 0)),
                            # total_debits=float(statement_data.get("total_debits", 0)),
                            # ending_balance=float(
                                # statement_data.get("ending_balance", 0)
                                # or statement_data.get("balance", 0)
                            # ),
                            # statement_date=statement_data.get("date_range", ""),
                            # details_json=json.dumps(payload, ensure_ascii=False),
                        # )
                        # db.session.add(s)
                        # db.session.commit()
                        # flash(f"Extrato importado: {ntx} transações.", "success")
                        # for w in warnings:
                            # flash(w, "info")
                        # return redirect(url_for("crm_statement"))
                # except Exception as e:
                    # flash(f"Erro ao processar CSV: {str(e)}", "danger")
        # summary = BankStatementSummary.query.order_by(BankStatementSummary.created_at.desc()).first()
        # if statement_data is None and summary:
            # statement_data = _display_statement_from_summary(summary)
        # return render_template(
            # "crm/statement.html",
            # statement_data=statement_data,
            # summary=summary,
            # last_updated=datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
        # )
    #
    @app.route("/crm/propertys")
    @app.route("/crm/propertys/regions-we")
    def crm_propertys():
        from types import SimpleNamespace

        regiao_filter = request.args.get("regiao", "") or ""
        community_filter = (request.args.get("community") or "").strip()
        bedrooms_f = (request.args.get("bedrooms") or "").strip()

        if regiao_filter == "W":
            comm_subset = [
                c[0]
                for c in db.session.query(RegionsWeUnit.community)
                .filter(RegionsWeUnit.region.in_(["W", "WE"]))
                .distinct()
                .all()
            ]
            q = RegionsWeUnit.query.filter(RegionsWeUnit.community.in_(comm_subset)) if comm_subset else RegionsWeUnit.query.filter(RegionsWeUnit.id == -1)
        elif regiao_filter == "E":
            comm_subset = [
                c[0]
                for c in db.session.query(RegionsWeUnit.community)
                .filter(RegionsWeUnit.region.in_(["E", "WE"]))
                .distinct()
                .all()
            ]
            q = RegionsWeUnit.query.filter(RegionsWeUnit.community.in_(comm_subset)) if comm_subset else RegionsWeUnit.query.filter(RegionsWeUnit.id == -1)
        else:
            q = RegionsWeUnit.query

        if community_filter:
            q = q.filter(RegionsWeUnit.community.ilike(f"%{community_filter}%"))
        if bedrooms_f:
            try:
                q = q.filter(RegionsWeUnit.bedrooms == int(bedrooms_f))
            except ValueError:
                pass

        units = q.all()
        agg = {}
        for u in units:
            comm = u.community
            if comm not in agg:
                agg[comm] = {"total_casas": 0, "total_quartos": 0, "regioes": set()}
            agg[comm]["total_casas"] += 1
            agg[comm]["total_quartos"] += int(u.bedrooms or 0)
            if u.region:
                agg[comm]["regioes"].add(u.region)

        table_rows = []
        for comm, d in sorted(agg.items(), key=lambda x: (-x[1]["total_casas"], x[0])):
            regs = sorted(d["regioes"])
            if "WE" in regs or ({"W", "E"}.issubset(set(regs))):
                regiao_str = "W, E" if "WE" not in regs else "WE"
            else:
                regiao_str = ", ".join(regs) if regs else ""
            table_rows.append(
                SimpleNamespace(
                    community=comm,
                    total_casas=d["total_casas"],
                    total_quartos=d["total_quartos"],
                    regiao=regiao_str,
                )
            )

        region_total_w = RegionsWeUnit.query.filter(RegionsWeUnit.region.in_(["W", "WE"])).count()
        region_total_e = RegionsWeUnit.query.filter(RegionsWeUnit.region.in_(["E", "WE"])).count()
        region_bedrooms_w = (
            db.session.query(db.func.coalesce(db.func.sum(RegionsWeUnit.bedrooms), 0))
            .filter(RegionsWeUnit.region.in_(["W", "WE"]))
            .scalar()
            or 0
        )
        region_bedrooms_e = (
            db.session.query(db.func.coalesce(db.func.sum(RegionsWeUnit.bedrooms), 0))
            .filter(RegionsWeUnit.region.in_(["E", "WE"]))
            .scalar()
            or 0
        )
        total_communities = db.session.query(RegionsWeUnit.community).distinct().count()
        properties_total_casas = RegionsWeUnit.query.count()
        properties_total_outs = RegionsWeUnit.query.filter_by(is_out=True).count()

        topn = sorted(agg.items(), key=lambda x: -x[1]["total_casas"])[:14]
        property_bar_labels = [t[0][:40] for t in topn]
        property_bar_values = [t[1]["total_casas"] for t in topn]

        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
        return render_template(
            "crm/regions_we.html",
            region_total_w=region_total_w,
            region_total_e=region_total_e,
            region_bedrooms_w=int(region_bedrooms_w),
            region_bedrooms_e=int(region_bedrooms_e),
            total_communities=total_communities,
            properties_total_casas=properties_total_casas,
            properties_total_outs=properties_total_outs,
            table_communities_distribution=table_rows,
            regiao_filter=regiao_filter,
            community_filter=community_filter,
            bedrooms_filter=bedrooms_f,
            property_bar_labels=property_bar_labels,
            property_bar_values=property_bar_values,
            last_updated=now,
        )

    @app.route("/crm/propertys/regions-we/upload", methods=["POST"])
    def crm_regions_we_upload():
        f = request.files.get("file")
        replace = request.form.get("replace_data") == "1"
        if not f or not f.filename:
            flash("Selecione um arquivo CSV ou Excel.", "warning")
            return redirect(url_for("crm_propertys"))
        ext = (f.filename or "").lower().rsplit(".", 1)[-1]
        rows = []
        errs = []
        try:
            if ext == "csv":
                raw = f.read()
                text = None
                for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
                    try:
                        text = raw.decode(enc)
                        break
                    except UnicodeDecodeError:
                        continue
                if text is None:
                    text = raw.decode("utf-8", errors="replace")
                delim = ";" if text.count(";") > text.count(",") else ","
                reader = csv.DictReader(io.StringIO(text), delimiter=delim)
                rows, errs = _regions_we_rows_from_dicts(reader.fieldnames, list(reader))
            elif ext == "xlsx":
                import openpyxl

                wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True)
                ws = wb.active
                it = ws.iter_rows(values_only=True)
                header = next(it, None)
                if not header:
                    flash("Planilha vazia.", "warning")
                    return redirect(url_for("crm_propertys"))
                headers = [str(c or "").strip() for c in header]
                dicts = []
                for row in it:
                    dicts.append(
                        {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
                    )
                rows, errs = _regions_we_rows_from_dicts(headers, dicts)
            else:
                flash("Formato inválido. Use .csv ou .xlsx", "danger")
                return redirect(url_for("crm_propertys"))
        except Exception as e:
            flash(f"Erro ao ler arquivo: {e}", "danger")
            return redirect(url_for("crm_propertys"))

        if errs:
            for e in errs:
                flash(e, "danger")
            return redirect(url_for("crm_propertys"))
        try:
            if replace:
                RegionsWeUnit.query.delete()
                db.session.commit()
            for r in rows:
                db.session.add(
                    RegionsWeUnit(
                        community=r["community"],
                        region=r["region"],
                        bedrooms=r["bedrooms"],
                        owner=r["owner"],
                        unit_name=r["unit_name"],
                        is_out=bool(r.get("is_out")),
                    )
                )
            db.session.commit()
            flash(f"{len(rows)} linha(s) importada(s) com sucesso.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao gravar na base: {e}", "danger")
        return redirect(url_for("crm_propertys"))

    @app.route("/crm/propertys/export-csv")
    def crm_propertys_export_csv():
        from flask import Response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["property", "region", "bedrooms", "owner", "status"])
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment; filename=propertys.csv"})

    @app.route("/crm/propertys/<int:prop_id>/toggle-out-hide", methods=["POST"])
    def crm_propertys_toggle_out_hide(prop_id):
        flash("Ocultar/mostrar em desenvolvimento.", "info")
        return redirect(url_for("crm_propertys"))

    @app.route("/crm/propertys/upload", methods=["POST"])
    def crm_propertys_upload():
        flash("Upload de propriedades em desenvolvimento.", "info")
        return redirect(url_for("crm_propertys"))

    @app.route("/crm/reservas")
    def crm_reservas():
        """Dashboard Reservations (CSV DataGrid no browser — static/reservations_agencies_dashboard.html)."""
        return render_template(
            "crm/reservas.html",
            last_updated=datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
        )

    @app.route("/crm/reservas/upload", methods=["POST"])
    def crm_reservas_upload():
        f = request.files.get("file")
        upload_mode = request.form.get("upload_mode", "integrar")
        if not f or f.filename == "":
            flash("Selecione um arquivo CSV.", "warning")
            return redirect(url_for("crm_reservas"))
        if not (f.filename or "").lower().endswith(".csv"):
            flash("Formato inválido. Use .csv", "danger")
            return redirect(url_for("crm_reservas"))
        try:
            content = f.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(content))
            headers = [h.strip() for h in (reader.fieldnames or [])]
            def _parse_float(v):
                if v is None or v == "": return 0.0
                s = str(v).strip().replace(",", "").replace(" ", "")
                s = re.sub(r"[^\d.\-]", "", s)
                try: return float(s) if s else 0.0
                except: return 0.0
            def _parse_date(v):
                if not v: return None
                s = str(v).strip()
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                    try: return datetime.strptime(s, fmt).date()
                    except ValueError: continue
                return None
            col_map = {"id": "reservation_id", "source": "source", "unit": "unit", "guest": "guest", "arrival date": "reservation_date",
                       "total (inc tax)": "total_inc_tax", "amount received": "amount_received", "property": "property_name", "owner": "owner"}
            n_processed = 0
            anos_set = set()
            propriedades_set = set()
            receita_total = 0.0
            meses = {}
            receita_por_mes = {}
            breakdown_fonte = {}
            top10_prop = {}
            if upload_mode == "substituir":
                Reserva.query.delete()
                db.session.commit()
            for r in reader:
                row = {k.strip().lower(): v for k, v in r.items() if k}
                unit = _normalize_source(row.get("unit") or row.get("property") or "")
                prop = _normalize_source(row.get("property") or row.get("unit") or unit)
                owner = _normalize_source(row.get("owner") or "")
                if not unit and not prop:
                    continue
                res_date = _parse_date(
                    row.get("arrival date")
                    or row.get("arrival_date")
                    or row.get("arrival")
                    or row.get("check-in")
                    or row.get("check in")
                    or row.get("checkin")
                    or row.get("start date")
                    or row.get("date")
                )
                rec = Reserva(
                    reservation_id=_normalize_source(row.get("id") or ""),
                    source=_normalize_source(row.get("source") or ""),
                    unit=unit or prop,
                    guest=_normalize_source(row.get("guest") or ""),
                    reservation_date=res_date,
                    nights=_parse_float(row.get("nights") or row.get("nights")),
                    total_inc_tax=_parse_float(row.get("total (inc tax)") or row.get("total_inc_tax") or row.get("total")),
                    amount_received=_parse_float(row.get("amount received") or row.get("amount_received") or row.get("total (inc tax)") or row.get("total")),
                    property_name=prop or unit,
                    owner=owner,
                )
                db.session.add(rec)
                n_processed += 1
                if res_date:
                    anos_set.add(res_date.year)
                    key = f"{res_date.year}-{res_date.month}"
                    meses[key] = meses.get(key, 0) + 1
                    receita_por_mes[key] = receita_por_mes.get(key, 0.0) + (rec.amount_received or 0)
                propriedades_set.add(prop or unit)
                receita_total += rec.amount_received or 0
                breakdown_fonte[rec.source or "N/A"] = breakdown_fonte.get(rec.source or "N/A", 0) + 1
                top10_prop[prop or unit] = top10_prop.get(prop or unit, 0) + 1
            db.session.commit()
            top10_list = sorted(top10_prop.items(), key=lambda x: -x[1])[:10]
            session["upload_summary"] = {
                "n_processed": n_processed,
                "anos": sorted(anos_set),
                "total_propriedades": len(propriedades_set),
                "receita_total": receita_total,
                "ticket_medio": receita_total / n_processed if n_processed else 0,
                "meses": meses,
                "receita_por_mes": receita_por_mes,
                "top10_propriedades": top10_list,
                "breakdown_fonte": breakdown_fonte,
                "alertas": [],
            }
            flash(f"{n_processed} reservas importadas com sucesso.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao importar: {e}", "danger")
        return redirect(url_for("crm_reservas"))

    @app.route("/crm/renovations")
    def crm_renovations():
        return _crm_stub("renovations", "crm/renovations.html", em_andamento=0, concluidas=0, atrasadas=0, total_invested=0, items=[])

    @app.route("/crm/renovations/export-csv")
    def crm_renovations_export_csv():
        from flask import Response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["project_name", "property_address", "start_date", "deadline", "total_value", "down_payment", "payments_made", "pct_complete", "status"])
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment; filename=renovations.csv"})

    @app.route("/crm/renovations/upload", methods=["POST"])
    def crm_renovations_upload():
        f = request.files.get("file")
        if not f or f.filename == "":
            flash("Selecione um arquivo CSV.", "warning")
            return redirect(url_for("crm_renovations"))
        flash("Upload de renovações em desenvolvimento. Use o formato: Projeto, Propriedade, Início, Prazo, Valor total, Entrada, Valor pago, Status.", "info")
        return redirect(url_for("crm_renovations"))

    @app.route("/crm/renovations/import-financial", methods=["POST"])
    def crm_renovations_import_financial():
        f = request.files.get("file")
        if not f or f.filename == "":
            flash("Selecione um arquivo CSV.", "warning")
            return redirect(url_for("crm_renovations"))
        flash("Importação financeira em desenvolvimento. Colunas: Data, Casa, Termos, Proprietário, Valor total, Valor devido.", "info")
        return redirect(url_for("crm_renovations"))

    @app.route("/crm/cleaners")
    def crm_cleaners():
        return render_template(
            "crm/cleaners.html",
            last_updated=datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
        )

    @app.route("/crm/cleaners/export-csv")
    def crm_cleaners_export_csv():
        from flask import Response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["company_name", "cleaner_name", "total_quantity", "total_value"])
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment; filename=cleaners.csv"})

    @app.route("/crm/cleaners/upload-plan-a", methods=["POST"])
    def crm_cleaners_upload_plan_a():
        flash("Planilha Housekeeper enviada. Módulo em desenvolvimento.", "info")
        return redirect(url_for("crm_cleaners"))

    @app.route("/crm/cleaners/upload-plan-b", methods=["POST"])
    def crm_cleaners_upload_plan_b():
        flash("Tabela com Nome das Empresas enviada. Módulo em desenvolvimento.", "info")
        return redirect(url_for("crm_cleaners"))

    @app.route("/crm/cleaners/merge", methods=["POST"])
    def crm_cleaners_merge():
        flash("Fusão em desenvolvimento. Envie plan_a e plan_b primeiro.", "info")
        return redirect(url_for("crm_cleaners"))

    @app.route("/crm/cleaners/upload-totals", methods=["POST"])
    def crm_cleaners_upload_totals():
        flash("Upload de totais em desenvolvimento.", "info")
        return redirect(url_for("crm_cleaners"))

    @app.route("/crm/cleaners/upload", methods=["POST"])
    def crm_cleaners_upload():
        flash("Upload de Cleaners em desenvolvimento. Colunas: company_name, cleaner_name, cleaning_date, Clean Cost to MC.", "info")
        return redirect(url_for("crm_cleaners"))

    @app.route("/crm/cleaners/merged/<int:row_id>/details")
    def crm_cleaners_merged_details(row_id):
        return jsonify({"total_quantity": 0, "total_value": 0, "details": []})

    @app.route("/crm/cleaners/merged/<int:row_id>/approve", methods=["POST"])
    def crm_cleaners_merged_approve(row_id):
        flash("Aprovação em desenvolvimento.", "info")
        return redirect(url_for("crm_cleaners"))

    @app.route("/crm/cleaners/merged/<int:row_id>/visited", methods=["POST"])
    def crm_cleaners_merged_visited(row_id):
        flash("Visitado em desenvolvimento.", "info")
        return redirect(url_for("crm_cleaners"))

    @app.route("/crm/contractors")
    def crm_contractors():
        csv_rows = []
        return _crm_stub(
            "contractors",
            "crm/contractors.html",
            contractors=[],
            csv_rows=csv_rows,
            contractor_empresas_count=_contractors_unique_company_count(csv_rows),
            last_csv_upload=None,
            month=datetime.utcnow().strftime("%Y-%m"),
            _stat_cost=0,
        )

    @app.route("/crm/contractors/export-csv")
    def crm_contractors_export_csv():
        from flask import Response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["contractor_name", "total_cost"])
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment; filename=contractors.csv"})

    @app.route("/crm/contractors/upload", methods=["POST"])
    def crm_contractors_upload():
        flash("Upload de contractors em desenvolvimento.", "info")
        return redirect(url_for("crm_contractors"))

    @app.route("/crm/payroll")
    def crm_payroll():
        payroll_rows = []
        employees = []
        month = request.args.get("month") or datetime.utcnow().strftime("%Y-%m")
        return _crm_stub(
            "payroll",
            "crm/payroll.html",
            payroll_rows=payroll_rows,
            employees=employees,
            emp_totals={},
            payroll_empresas_count=_payroll_distinct_department_count(payroll_rows, employees),
            payroll_dept_labels=["Operações", "Financeiro", "Housekeeping", "Manutenção"],
            payroll_headcount=[0, 0, 0, 0],
            payroll_costs=[0.0, 0.0, 0.0, 0.0],
            total_payroll=0.0,
            total_hours=0.0,
            month=month,
        )

    @app.route("/crm/payroll/upload", methods=["POST"])
    def crm_payroll_upload():
        flash("Payroll CSV recebido (integração com base em desenvolvimento).", "info")
        m = request.form.get("month") or datetime.utcnow().strftime("%Y-%m")
        return redirect(url_for("crm_payroll", month=m))

    @app.route("/crm/payroll/bulk-delete", methods=["POST"])
    def crm_payroll_bulk_delete():
        flash("Exclusão em lote em desenvolvimento.", "info")
        m = request.form.get("month") or datetime.utcnow().strftime("%Y-%m")
        return redirect(url_for("crm_payroll", month=m))

    @app.route("/crm/payroll/export-csv")
    def crm_payroll_export_csv():
        from flask import Response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["employee", "amount", "period"])
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment; filename=payroll.csv"})

    @app.route("/crm/licenses")
    def crm_licenses():
        return _crm_stub("licenses", "crm/licenses.html", licenses=[], items=[], lic_type="DPBR", type_counts={}, city_county_counts={}, miami_count=0, office_filter=None, city_county_filter=None, em_dia=0, a_vencer=0, vencidas=0)

    @app.route("/crm/licenses/export-csv")
    def crm_licenses_export_csv():
        from flask import Response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["license_number", "city_county", "issue_date", "expiry_date", "status", "property_ref"])
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment; filename=licenses.csv"})

    @app.route("/crm/licenses/upload", methods=["POST"])
    def crm_licenses_upload():
        flash("Upload de licenças em desenvolvimento. Colunas: license_number, city_county, issue_date, expiry_date.", "info")
        return redirect(url_for("crm_licenses"))

    @app.route("/crm/licenses/bulk-delete", methods=["POST"])
    def crm_licenses_bulk_delete():
        flash("Exclusão em lote em desenvolvimento.", "info")
        return redirect(url_for("crm_licenses"))

    @app.route("/crm/licenses/<int:license_id>/mark-renewed", methods=["POST"])
    def crm_licenses_mark_renewed(license_id):
        flash("Marcar renovada em desenvolvimento.", "info")
        return redirect(url_for("crm_licenses"))

    @app.route("/crm/licenses/<int:license_id>/edit")
    def crm_licenses_edit(license_id):
        flash("Edição em desenvolvimento.", "info")
        return redirect(url_for("crm_licenses"))

    @app.route("/crm/stock")
    def crm_stock():
        month_names = {i: ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"][i-1] for i in range(1, 13)}
        return _crm_stub("stock", "crm/stock.html", offices=["Office 1", "Office 2", "Office 3", "Miami"], counts={}, total_gasto=0, month=datetime.utcnow().month, year=datetime.utcnow().year, month_names=month_names, selected_office=None, items=[])

    @app.route("/crm/stock/upload", methods=["POST"])
    def crm_stock_upload():
        flash("Upload de stock em desenvolvimento.", "info")
        return redirect(url_for("crm_stock"))

    # --- [rota desativada — menu removido] ---
    # @app.route("/crm/divvy")
    # def crm_divvy():
        # return _crm_stub("divvy", "crm/divvy.html", total=0, expenses=[], items=[], hidden_rows=[], dept_buttons=[], month=datetime.utcnow().strftime("%Y-%m"), dept_filter="", status_filter="", dash_subtitle="", total_count=0, search_q="")
    #
    # @app.route("/crm/divvy/export-csv")
    # def crm_divvy_export_csv():
        # from flask import Response
        # output = io.StringIO()
        # w = csv.writer(output)
        # w.writerow(["date", "category", "merchant", "amount", "cardholder", "department"])
        # return Response(output.getvalue(), mimetype="text/csv",
                       # headers={"Content-Disposition": "attachment; filename=divvy.csv"})
    #
    # @app.route("/crm/divvy/export-pdf")
    # def crm_divvy_export_pdf():
        # flash("Exportação PDF em desenvolvimento.", "info")
        # return redirect(url_for("crm_divvy"))
    #
    # @app.route("/crm/divvy/upload", methods=["POST"])
    # def crm_divvy_upload():
        # flash("Upload Divvy em desenvolvimento.", "info")
        # return redirect(url_for("crm_divvy"))
    #
    # @app.route("/crm/divvy/<int:expense_id>/toggle-hide", methods=["POST"])
    # def crm_divvy_toggle_hide(expense_id):
        # flash("Ocultar/mostrar em desenvolvimento.", "info")
        # return redirect(url_for("crm_divvy"))
    #
    # @app.route("/crm/monthlog")
    # def crm_monthlog():
        # return _crm_stub("monthlog", "crm/monthlog.html", rows=[])
    #
    # @app.route("/crm/monthlog/export-csv")
    # def crm_monthlog_export_csv():
        # from flask import Response
        # output = io.StringIO()
        # w = csv.writer(output)
        # w.writerow(["month", "year", "description", "value"])
        # return Response(output.getvalue(), mimetype="text/csv",
                       # headers={"Content-Disposition": "attachment; filename=monthlog.csv"})
    #
    # @app.route("/crm/monthlog/upload", methods=["POST"])
    # def crm_monthlog_upload():
        # flash("Upload de monthlog em desenvolvimento.", "info")
        # return redirect(url_for("crm_monthlog"))
    #
    @app.route("/crm/cartoes")
    def crm_cartoes():
        return _crm_stub("cartoes", "crm/cartoes.html")

    @app.route("/crm/cartoes/export-csv")
    def crm_cartoes_export_csv():
        from flask import Response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["card_name", "limit", "balance"])
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment; filename=cartoes.csv"})

    @app.route("/crm/cartoes/upload", methods=["POST"])
    def crm_cartoes_upload():
        flash("Upload de cartões em desenvolvimento.", "info")
        return redirect(url_for("crm_cartoes"))

    @app.route("/crm/api/cars-insurance", methods=["GET", "POST"])
    def crm_api_cars_insurance():
        """Lê/grava snapshot JSON da frota + seguros (dedup por VIN no cliente)."""
        if request.method == "GET":
            snap = CarsInsuranceSnapshot.query.order_by(CarsInsuranceSnapshot.updated_at.desc()).first()
            if not snap:
                return jsonify({"ok": True, "rows": [], "updated_at": None})
            try:
                data = json.loads(snap.payload_json or "{}")
            except (json.JSONDecodeError, TypeError):
                data = {}
            rows = data.get("rows") if isinstance(data.get("rows"), list) else []
            ua = snap.updated_at.isoformat() + "Z" if snap.updated_at else None
            return jsonify({"ok": True, "rows": rows, "updated_at": ua})
        data = request.get_json(silent=True) or {}
        rows = data.get("rows")
        if not isinstance(rows, list):
            return jsonify({"ok": False, "error": "rows must be a list"}), 400
        if len(rows) > 5000:
            return jsonify({"ok": False, "error": "too many rows"}), 400
        def _ci_date_field(v):
            if v is None or v == "":
                return None
            s = str(v).strip()[:10]
            if len(s) >= 10 and s[4:5] == "-" and s[7:8] == "-":
                return s
            return None

        clean = []
        for r in rows:
            if not isinstance(r, dict):
                continue
            vin = str(r.get("vin") or "").strip()[:32]
            if not vin:
                continue
            try:
                val_num = float(r.get("valor") or 0)
            except (TypeError, ValueError):
                val_num = 0.0
            clean.append(
                {
                    "vin": vin,
                    "veiculo": str(r.get("veiculo") or "")[:500],
                    "year": str(r.get("year") or "")[:8],
                    "make": str(r.get("make") or "")[:120],
                    "model": str(r.get("model") or "")[:200],
                    "plate": str(r.get("plate") or "")[:32],
                    "insurer": str(r.get("insurer") or "")[:200],
                    "insDate": _ci_date_field(r.get("insDate")),
                    "compraDate": _ci_date_field(r.get("compraDate")),
                    "valor": val_num,
                }
            )
            if len(clean) > 5000:
                break
        try:
            CarsInsuranceSnapshot.query.delete()
            db.session.add(
                CarsInsuranceSnapshot(
                    payload_json=json.dumps({"rows": clean}, ensure_ascii=False),
                )
            )
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"ok": False, "error": str(e)}), 500
        return jsonify({"ok": True, "saved": len(clean)})

    @app.route("/crm/cars")
    def crm_cars():
        return render_template("crm/cars.html", last_updated=datetime.utcnow().strftime("%Y-%m-%d %H:%M"))

    @app.route("/crm/cars/export-csv")
    def crm_cars_export_csv():
        from flask import Response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["vin", "make", "model", "year", "plate"])
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment; filename=cars.csv"})

    @app.route("/crm/cars/bulk-delete", methods=["POST"])
    def crm_cars_bulk_delete():
        flash("Exclusão em lote em desenvolvimento.", "info")
        return redirect(url_for("crm_cars"))

    @app.route("/crm/cars/upload", methods=["POST"])
    def crm_cars_upload():
        flash("Upload de veículos em desenvolvimento.", "info")
        return redirect(url_for("crm_cars"))

    @app.route("/crm/cars/<int:car_id>/edit")
    def crm_cars_edit(car_id):
        flash("Edição em desenvolvimento.", "info")
        return redirect(url_for("crm_cars"))

    @app.route("/crm/cars/<int:car_id>/delete", methods=["POST"])
    def crm_cars_delete(car_id):
        flash("Exclusão em desenvolvimento.", "info")
        return redirect(url_for("crm_cars"))

    @app.route("/crm/insurance")
    def crm_insurance():
        return render_template("crm/insurance.html", last_updated=datetime.utcnow().strftime("%Y-%m-%d %H:%M"))


    @app.route("/crm/insurance/export-csv")
    def crm_insurance_export_csv():
        from flask import Response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["vin", "insurer", "policy", "expiry"])
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment; filename=insurance.csv"})

    @app.route("/crm/insurance/toggle-hide-card", methods=["POST"])
    def crm_insurance_toggle_hide_card():
        flash("Ocultar/mostrar em desenvolvimento.", "info")
        return redirect(url_for("crm_insurance"))

    @app.route("/crm/insurance/delete-selected", methods=["POST"])
    def crm_insurance_delete_selected():
        flash("Exclusão selecionada em desenvolvimento.", "info")
        return redirect(url_for("crm_insurance"))

    @app.route("/crm/insurance/upload", methods=["POST"])
    def crm_insurance_upload():
        flash("Upload de seguros em desenvolvimento.", "info")
        return redirect(url_for("crm_insurance"))

    @app.route("/crm/long-term")
    @app.route("/crm/long_term")
    def crm_long_term():
        snap = LongTermRollSnapshot.query.order_by(LongTermRollSnapshot.updated_at.desc()).first()
        roll: dict = {}
        if snap and snap.payload_json:
            try:
                roll = json.loads(snap.payload_json)
            except (json.JSONDecodeError, TypeError):
                roll = {}
        actions_map = {a.row_key: a.action for a in LongTermRollRowAction.query.all()}
        roll_logs = (
            LongTermRollUploadLog.query.order_by(LongTermRollUploadLog.created_at.desc()).limit(40).all()
        )
        receita_mensal = 0.0
        ocupadas = vagas = ativos = total_casas = 0
        if roll and roll.get("kpis"):
            k = roll["kpis"]
            receita_mensal = float(k.get("total_rent_monthly") or 0)
            ocupadas = int(k.get("casas_alugadas") or 0)
            vagas = int(k.get("casas_vagas") or 0)
            total_casas = int(k.get("total_casas") or 0)
            ativos = ocupadas + vagas
        return render_template(
            "crm/long_term.html",
            last_updated=datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
            receita_mensal=receita_mensal,
            inadimplencia=float(roll["kpis"].get("total_past_due") or 0) if roll and roll.get("kpis") else 0,
            repasse=0,
            contracts=[],
            items=[],
            total_casas=total_casas,
            ativos=ativos or total_casas,
            ocupadas=ocupadas,
            vagas=vagas,
            subgrupos=[],
            sub_labels={},
            sub_totals={},
            sub=None,
            roll=roll,
            roll_actions_map=actions_map,
            roll_logs=roll_logs,
        )

    @app.route("/crm/long-term/rent-roll/upload", methods=["POST"])
    def crm_long_term_rent_roll_upload():
        from website.crm_long_term_roll import process_upload_bundle, recompute_full_payload

        mode = (request.form.get("upload_mode") or "replace").strip().lower()
        if mode not in ("replace", "append"):
            mode = "replace"

        def _take(name):
            f = request.files.get(name)
            if f and f.filename:
                return f.read()
            return None

        files = {}
        fnames = []
        mapping = [
            ("rent_roll", "rent_roll_csv"),
            ("tenant_directory", "tenant_directory_csv"),
            ("unit_directory", "unit_directory_csv"),
            ("property_directory", "property_directory_csv"),
            ("bill_detail", "bill_detail_csv"),
            ("owner_directory", "owner_directory_csv"),
        ]
        for key, field in mapping:
            raw = _take(field)
            if raw:
                files[key] = raw
                fnames.append(f"{key}:{request.files[field].filename[:80]}")

        if "tenant_directory" not in files:
            flash("Envie tenant_directory.csv (obrigatório — fonte principal de inquilinos).", "warning")
            return redirect(url_for("crm_long_term"))

        new_payload, errs = process_upload_bundle(files)
        if mode == "append":
            old = LongTermRollSnapshot.query.order_by(LongTermRollSnapshot.updated_at.desc()).first()
            if old and old.payload_json:
                try:
                    old_d = json.loads(old.payload_json)
                    old_meta = old_d.get("meta") or {}
                    byk = {r["row_key"]: r for r in old_d.get("rows", [])}
                    for r in new_payload.get("rows", []):
                        prev = byk.get(r["row_key"])
                        if prev:
                            if (not r.get("last_payment_date")) and prev.get("last_payment_date"):
                                r["last_payment_date"] = prev.get("last_payment_date")
                                r["last_payment_label"] = prev.get("last_payment_label")
                        byk[r["row_key"]] = r
                    merged_meta = {**old_meta, **(new_payload.get("meta") or {})}
                    nm = new_payload.get("meta") or {}
                    if float(nm.get("sum_units_property_dir") or 0) > 0:
                        merged_meta["sum_units_property_dir"] = nm["sum_units_property_dir"]
                    if int(nm.get("unique_properties_count") or 0) > 0:
                        merged_meta["unique_properties_count"] = nm["unique_properties_count"]
                    if nm.get("rent_roll_past_due_total") is not None:
                        merged_meta["rent_roll_past_due_total"] = nm["rent_roll_past_due_total"]
                    if nm.get("global_last_paid_iso"):
                        merged_meta["global_last_paid_iso"] = nm["global_last_paid_iso"]
                    if isinstance(nm.get("line"), dict):
                        merged_meta["line"] = nm["line"]
                    new_payload = recompute_full_payload(list(byk.values()), meta=merged_meta)
                except (json.JSONDecodeError, TypeError, KeyError):
                    pass

        if errs and new_payload.get("meta"):
            new_payload["meta"]["errors"] = errs

        try:
            LongTermRollSnapshot.query.delete()
            db.session.add(
                LongTermRollSnapshot(
                    payload_json=json.dumps(new_payload, ensure_ascii=False),
                )
            )
            user_lbl = _lt_actor_label()
            db.session.add(
                LongTermRollUploadLog(
                    mode=mode,
                    user_label=user_lbl,
                    filenames=" | ".join(fnames)[:580],
                    rows_imported=len(new_payload.get("rows") or []),
                    validation_ok=not bool(errs),
                    message=("; ".join(errs))[:500] if errs else "OK",
                )
            )
            db.session.commit()
            flash(
                "Upload concluído. Dados normalizados. Dashboard atualizado com KPIs e ações por linha.",
                "success",
            )
            if errs:
                for e in errs[:5]:
                    flash(e, "info")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao gravar rent roll: {e}", "danger")
        return redirect(url_for("crm_long_term"))

    @app.route("/crm/long-term/rent-roll/row-action", methods=["POST"])
    def crm_long_term_rent_roll_row_action():
        data = request.get_json(silent=True) or {}
        rk = (data.get("row_key") or "")[:64]
        act = (data.get("action") or "").lower().strip()
        if not rk or act not in ("approved", "declined", "review", "reset"):
            return jsonify({"ok": False}), 400
        user_lbl = _lt_actor_label()
        try:
            if act == "reset":
                LongTermRollRowAction.query.filter_by(row_key=rk).delete()
            else:
                row = LongTermRollRowAction.query.filter_by(row_key=rk).first()
                if row:
                    row.action = act
                    row.user_label = user_lbl
                else:
                    db.session.add(LongTermRollRowAction(row_key=rk, action=act, user_label=user_lbl))
                db.session.add(
                    LongTermRollAuditLog(row_key=rk, action=act, user_label=user_lbl)
                )
            db.session.commit()
            return jsonify({"ok": True})
        except Exception:
            db.session.rollback()
            return jsonify({"ok": False}), 500

    @app.route("/crm/long-term/rent-roll/export.csv")
    def crm_long_term_rent_roll_export():
        snap = LongTermRollSnapshot.query.order_by(LongTermRollSnapshot.updated_at.desc()).first()
        if not snap or not snap.payload_json:
            flash("Sem dados de rent roll para exportar.", "warning")
            return redirect(url_for("crm_long_term"))
        try:
            data = json.loads(snap.payload_json)
        except (json.JSONDecodeError, TypeError):
            flash("Snapshot inválido.", "danger")
            return redirect(url_for("crm_long_term"))
        rows = data.get("rows") or []
        actions = {a.row_key: a.action for a in LongTermRollRowAction.query.all()}
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(
            [
                "Property",
                "Unit",
                "Tenant",
                "Status",
                "Rent",
                "Deposit",
                "Market_Rent",
                "Lease_From",
                "Lease_To",
                "LAST_PAYMENT_DATE",
                "Past_Due",
                "Late_Count",
                "NSF_Count",
                "Action",
            ]
        )
        for r in rows:
            w.writerow(
                [
                    r.get("property_full", ""),
                    r.get("unit", ""),
                    r.get("tenant", ""),
                    r.get("status", ""),
                    r.get("rent", ""),
                    r.get("deposit", ""),
                    r.get("market_rent", ""),
                    r.get("lease_from", ""),
                    r.get("lease_to", ""),
                    r.get("last_payment_label", ""),
                    r.get("past_due", ""),
                    r.get("late_count", ""),
                    r.get("nsf_count", ""),
                    actions.get(r.get("row_key", ""), ""),
                ]
            )
        return Response(
            output.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=long_term_rent_roll_export.csv"},
        )

    @app.route("/crm/long-term/export-csv")
    def crm_long_term_export_csv():
        from flask import Response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["address", "owner", "status", "monthly_rent"])
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment; filename=long_term.csv"})

    @app.route("/crm/long-term/remove-duplicates", methods=["POST"])
    def crm_long_term_remove_duplicates():
        flash("Remoção de duplicados em desenvolvimento.", "info")
        return redirect(url_for("crm_long_term"))

    @app.route("/crm/long-term/upload", methods=["POST"])
    def crm_long_term_upload():
        flash("Upload de contratos em desenvolvimento.", "info")
        return redirect(url_for("crm_long_term"))

    @app.route("/crm/long-term/upload-account-totals", methods=["POST"])
    def crm_long_term_upload_account_totals():
        flash("Upload de totais em desenvolvimento.", "info")
        return redirect(url_for("crm_long_term"))

    @app.route("/crm/long-term/upload-owner-directory", methods=["POST"])
    def crm_long_term_upload_owner_directory():
        flash("Upload de diretório de owners em desenvolvimento.", "info")
        return redirect(url_for("crm_long_term"))

    @app.route("/crm/long-term/upload-appsheet", methods=["POST"])
    def crm_long_term_upload_appsheet():
        flash("Upload AppSheet em desenvolvimento.", "info")
        return redirect(url_for("crm_long_term"))

    @app.route("/crm/long-term/bulk", methods=["POST"])
    def crm_long_term_bulk():
        flash("Exclusão em lote em desenvolvimento.", "info")
        return redirect(url_for("crm_long_term"))

    @app.route("/crm/long-term/<int:unit_id>/toggle-hide", methods=["POST"])
    def crm_long_term_toggle_hide(unit_id):
        flash("Ocultar/mostrar em desenvolvimento.", "info")
        return redirect(url_for("crm_long_term"))

    @app.route("/crm/agencias")
    def crm_agencias():
        return render_template(
            "crm/agencias.html",
            last_updated=datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
        )

    @app.route("/crm/agencias/upload", methods=["POST"])
    def crm_agencias_upload():
        flash("Upload de agências em desenvolvimento.", "info")
        return redirect(url_for("crm_agencias"))

    @app.route("/crm/vendor")
    def crm_vendor():
        items = []
        ativos, inativos = 0, 0
        vendor_empresas_count = len(items) if items else (ativos + inativos)
        return _crm_stub(
            "vendor",
            "crm/vendor.html",
            ativos=ativos,
            inativos=inativos,
            total_pago=0,
            vendors=[],
            items=items,
            payments=[],
            vendor_empresas_count=vendor_empresas_count,
        )

    @app.route("/crm/vendor/payments/upload", methods=["POST"])
    def crm_vendor_payments_upload():
        flash("Upload de pagamentos de vendor em desenvolvimento.", "info")
        return redirect(url_for("crm_vendor"))

    @app.route("/crm/vendor/upload", methods=["POST"])
    def crm_vendor_upload():
        flash("Upload de vendors em desenvolvimento.", "info")
        return redirect(url_for("crm_vendor"))

    @app.route("/crm/cofre")
    def crm_cofre():
        return _crm_stub("cofre", "crm/cofre.html", balance=0, total_entradas=0, total_saidas=0, total_depositos=0, movements=[])

    @app.route("/crm/cofre/export-csv")
    def crm_cofre_export_csv():
        from flask import Response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["date", "type", "amount", "description"])
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment; filename=cofre_auditoria.csv"})

    @app.route("/crm/cofre/entrada", methods=["POST"])
    def crm_cofre_entrada():
        flash("Entrada em desenvolvimento.", "info")
        return redirect(url_for("crm_cofre"))

    @app.route("/crm/cofre/saida", methods=["POST"])
    def crm_cofre_saida():
        flash("Saída em desenvolvimento.", "info")
        return redirect(url_for("crm_cofre"))

    @app.route("/crm/cofre/deposito", methods=["POST"])
    def crm_cofre_deposito():
        flash("Depósito em desenvolvimento.", "info")
        return redirect(url_for("crm_cofre"))

    @app.route("/crm/cofre/<int:mid>/corrigir", methods=["POST"])
    def crm_cofre_corrigir(mid):
        flash("Correção em desenvolvimento.", "info")
        return redirect(url_for("crm_cofre"))

    # --- [rota desativada — menu removido] ---
    # @app.route("/crm/dp-plus")
    # @app.route("/crm/dp_plus")
    # def crm_dp_plus():
        # return render_template(
            # "crm/dp_plus.html",
            # last_updated=datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
        # )
    #
    # @app.route("/crm/dp-plus/<int:entry_id>/toggle-hide", methods=["POST"])
    # def crm_dp_plus_toggle_hide(entry_id):
        # flash("Ocultar/mostrar em desenvolvimento.", "info")
        # return redirect(url_for("crm_dp_plus"))
    #
    # @app.route("/crm/dp-plus/add", methods=["POST"])
    # def crm_dp_plus_add():
        # flash("Adicionar em desenvolvimento.", "info")
        # return redirect(url_for("crm_dp_plus"))
    #
    @app.route("/crm/upcoming")
    def crm_upcoming():
        return render_template(
            "crm/upcoming.html",
            last_updated=datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
        )

    @app.route("/crm/upcoming/upload", methods=["POST"])
    def crm_upcoming_upload():
        flash("Upload de upcoming em desenvolvimento.", "info")
        return redirect(url_for("crm_upcoming"))

    @app.route("/crm/upcoming/<int:row_id>/toggle-hide", methods=["POST"])
    def crm_upcoming_toggle_hide(row_id):
        flash("Ocultar/mostrar em desenvolvimento.", "info")
        return redirect(url_for("crm_upcoming"))

    @app.route("/crm/1099")
    def crm_1099():
        return render_template("crm/1099.html")

    @app.route("/crm/integrations")
    def crm_integrations_page():
        connections = IntegrationConnection.query.order_by(
            IntegrationConnection.provider, IntegrationConnection.id
        ).all()
        logs = IntegrationLog.query.order_by(IntegrationLog.created_at.desc()).limit(100).all()
        return render_template(
            "crm/integrations.html",
            connections=connections,
            logs=logs,
            integration_providers=INTEGRATION_PROVIDERS,
        )

    @app.route("/api/integrations/status")
    def api_integrations_status():
        """JSON para dashboards: conexões e último sync."""
        out = []
        for c in IntegrationConnection.query.order_by(IntegrationConnection.provider).all():
            out.append(
                {
                    "id": c.id,
                    "provider": c.provider,
                    "label": c.label,
                    "external_account_id": c.external_account_id,
                    "status": c.status,
                    "last_sync_at": c.last_sync_at.isoformat() + "Z" if c.last_sync_at else None,
                    "last_error": c.last_error,
                }
            )
        return jsonify({"ok": True, "connections": out})

    @app.route("/api/ramp/transactions")
    def api_ramp_transactions():
        client_id = os.environ.get("RAMP_CLIENT_ID")
        client_secret = os.environ.get("RAMP_CLIENT_SECRET")
        if not client_id or not client_secret:
            return jsonify({"error": "Ramp credentials not configured"}), 500
        creds = base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()
        token_res = req.post(
            "https://api.ramp.com/developer/v1/token",
            headers={
                "Authorization": f"Basic {creds}",
                "Content-Type": "application/x-www-form-urlencoded",
            },
            data={"grant_type": "client_credentials", "scope": "transactions:read"},
        )
        token = token_res.json().get("access_token")
        params = {"page_size": 100}
        from_date = request.args.get("from_date")
        if from_date:
            params["from_date"] = from_date
        tx_res = req.get(
            "https://api.ramp.com/developer/v1/transactions",
            headers={"Authorization": f"Bearer {token}"},
            params=params,
        )
        return jsonify(tx_res.json())

    @app.route("/api/ramp/cards")
    def api_ramp_cards():
        client_id = os.environ.get("RAMP_CLIENT_ID")
        client_secret = os.environ.get("RAMP_CLIENT_SECRET")
        if not client_id or not client_secret:
            return jsonify({"error": "Ramp credentials not configured"}), 500
        creds = base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()
        token_res = req.post(
            "https://api.ramp.com/developer/v1/token",
            headers={
                "Authorization": f"Basic {creds}",
                "Content-Type": "application/x-www-form-urlencoded",
            },
            data={"grant_type": "client_credentials", "scope": "cards:read"},
        )
        token = token_res.json().get("access_token")
        cr_res = req.get(
            "https://api.ramp.com/developer/v1/cards",
            headers={"Authorization": f"Bearer {token}"},
            params={"page_size": 100},
        )
        return jsonify(cr_res.json())

    @app.route("/api/webhooks/<provider>", methods=["POST"])
    def api_integration_webhook(provider):
        p = (provider or "").lower().strip()
        allowed = {x[0] for x in INTEGRATION_PROVIDERS}
        if p not in allowed:
            return jsonify({"ok": False, "error": "unknown provider"}), 404
        raw = request.get_data(cache=False) or b""
        hdrs = {k: v for k, v in request.headers.items()}
        sig_ok = integration_verify_webhook_signature(p, raw, hdrs)
        preview = raw[:500].decode("utf-8", errors="replace")
        ev = request.headers.get("X-Request-Id") or request.headers.get("X-Event-Id") or ""
        db.session.add(
            IntegrationWebhookDelivery(
                provider=p,
                event_id=(ev or "")[:120],
                signature_ok=sig_ok,
                body_preview=preview,
            )
        )
        log = IntegrationLog(
            provider=p,
            action="webhook_in",
            status="ok" if sig_ok else "error",
            message="webhook received",
            details_json=json.dumps({"signature_ok": sig_ok, "bytes": len(raw)}),
        )
        db.session.add(log)
        db.session.commit()
        if not sig_ok:
            return jsonify({"ok": False, "error": "invalid signature"}), 401
        return jsonify({"ok": True, "received": True})

    @app.route("/api/integrations/oauth/<provider>/start")
    def api_integration_oauth_start(provider):
        p = (provider or "").lower().strip()
        allowed = {x[0] for x in INTEGRATION_PROVIDERS}
        if p not in allowed:
            flash("Provider desconhecido.", "danger")
            return redirect(url_for("crm_integrations_page"))
        flash(
            "OAuth: configure client_id/secret em variáveis de ambiente e implemente o redirect do provider "
            + p
            + ".",
            "info",
        )
        return redirect(url_for("crm_integrations_page"))

    @app.route("/api/integrations/oauth/<provider>/callback")
    def api_integration_oauth_callback(provider):
        p = (provider or "").lower().strip()
        flash("Callback OAuth recebido para " + p + " — persistir tokens com integration_encrypt_token.", "info")
        return redirect(url_for("crm_integrations_page"))

    @app.route("/api/integrations/sync/<int:connection_id>", methods=["POST"])
    def api_integration_enqueue_sync(connection_id):
        """Enfileira sync — Celery pode consumir integration_sync_jobs; sem worker fica pending."""
        c = IntegrationConnection.query.get_or_404(connection_id)
        job = IntegrationSyncJob(connection_id=c.id, job_type="full_sync", status="pending")
        db.session.add(job)
        db.session.add(
            IntegrationLog(
                connection_id=c.id,
                provider=c.provider,
                action="sync_enqueued",
                status="ok",
                message="Job criado (pending)",
            )
        )
        db.session.commit()
        return jsonify({"ok": True, "job_id": job.id})

    @app.route("/crm/news", methods=["GET", "POST"])
    def crm_news():
        _news_session_user_key()
        if request.method == "POST":
            body = (request.form.get("body") or "").strip()
            cat = (request.form.get("category") or "Geral").strip()[:32]
            auth = (request.form.get("author_name") or _news_reader_display_name())[:80]
            role = (request.form.get("author_role") or "")[:80]
            allowed_cat = {"Invoice", "Payroll", "Statement", "Urgente", "Geral"}
            if cat not in allowed_cat:
                cat = "Geral"
            if len(body) < 1 or len(body) > 60:
                flash("A mensagem deve ter entre 1 e 60 caracteres.", "warning")
            else:
                db.session.add(
                    NewsPost(body=body, category=cat, author_name=auth, author_role=role)
                )
                db.session.commit()
                flash("Atualização publicada.", "success")
            return redirect(url_for("crm_news"))

        if NewsPost.query.count() == 0:
            db.session.add(
                NewsPost(
                    body="Bem-vindo ao feed Master Finance — máx. 60 caracteres.",
                    author_name="Master Finance",
                    author_role="Sistema",
                    category="Geral",
                )
            )
            db.session.commit()

        posts = NewsPost.query.order_by(NewsPost.created_at.desc()).limit(120).all()
        user_key = _news_session_user_key()
        read_ids = {
            r.post_id
            for r in NewsReadReceipt.query.filter(NewsReadReceipt.user_key == user_key).all()
        }
        enriched = []
        for p in posts:
            rc = NewsReadReceipt.query.filter_by(post_id=p.id).count()
            reacts = (
                db.session.query(NewsReaction.emoji, db.func.count(NewsReaction.id))
                .filter(NewsReaction.post_id == p.id)
                .group_by(NewsReaction.emoji)
                .all()
            )
            my_reacts = {
                r.emoji
                for r in NewsReaction.query.filter_by(post_id=p.id, user_key=user_key).all()
            }
            enriched.append(
                {
                    "post": p,
                    "read_count": rc,
                    "i_read": p.id in read_ids,
                    "reactions": [{"emoji": e, "count": c} for e, c in reacts],
                    "my_reactions": my_reacts,
                }
            )
        unread = sum(1 for x in enriched if not x["i_read"])
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
        pl_ctx = _process_library_template_context()
        return render_template(
            "crm/news.html",
            posts_enriched=enriched,
            news_unread_count=unread,
            last_updated=now,
            **pl_ctx,
        )

    @app.route("/crm/process-library/file/<int:pid>")
    def crm_process_library_file(pid):
        proc = ProcessLibraryItem.query.get_or_404(pid)
        base = current_app.config["PROCESS_LIBRARY_STORAGE"]
        return send_from_directory(
            base,
            proc.stored_filename,
            mimetype="application/pdf",
            as_attachment=False,
            download_name=proc.original_filename or "documento.pdf",
        )

    @app.route("/crm/process-library/download/<int:pid>")
    def crm_process_library_download(pid):
        proc = ProcessLibraryItem.query.get_or_404(pid)
        base = current_app.config["PROCESS_LIBRARY_STORAGE"]
        return send_from_directory(
            base,
            proc.stored_filename,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=proc.original_filename or "documento.pdf",
        )

    @app.route("/crm/news/api/process-library/summary", methods=["GET"])
    def crm_process_library_api_summary():
        return jsonify(_process_library_api_summary_dict())

    @app.route("/crm/process-library/upload", methods=["POST"])
    def crm_process_library_upload():
        if os.getenv("PROCESS_UPLOAD_REQUIRES_LOGIN", "").lower() in ("1", "true", "yes"):
            if not session.get("user_id"):
                return jsonify({"ok": False, "error": "É necessário iniciar sessão para enviar processos."}), 401
        name = (request.form.get("name") or "").strip()
        area = (request.form.get("area") or "").strip()
        version = (request.form.get("version") or "v1.0").strip()[:24]
        responsible = (request.form.get("responsible") or "").strip()[:120]
        desc = (request.form.get("description") or "").strip()[:120]
        f = request.files.get("file")
        if not name:
            return jsonify({"ok": False, "error": "Nome do processo é obrigatório."}), 400
        if area not in PROCESS_LIBRARY_AREAS:
            return jsonify({"ok": False, "error": "Área inválida."}), 400
        if not f or not f.filename:
            return jsonify({"ok": False, "error": "Selecione um arquivo PDF."}), 400
        if not f.filename.lower().endswith(".pdf"):
            return jsonify({"ok": False, "error": "Apenas ficheiros .pdf são permitidos."}), 400
        raw = f.read()
        if len(raw) > 20 * 1024 * 1024:
            return jsonify({"ok": False, "error": "Tamanho máximo: 20 MB."}), 400
        if len(raw) < 8 or raw[:4] != b"%PDF":
            return jsonify({"ok": False, "error": "O ficheiro não parece ser um PDF válido."}), 400
        uid = session.get("user_id")
        u = User.query.get(uid) if uid else None
        uploader = (
            (" ".join(x for x in (u.first_name, u.last_name) if x).strip() or u.username or "User")[:80]
            if u
            else _news_reader_display_name()
        )
        fn = uuid.uuid4().hex + ".pdf"
        path = os.path.join(current_app.config["PROCESS_LIBRARY_STORAGE"], fn)
        with open(path, "wb") as out:
            out.write(raw)
        proc = ProcessLibraryItem(
            name=name[:200],
            area=area,
            version=version,
            responsible=responsible,
            stored_filename=fn,
            original_filename=(f.filename or "documento.pdf")[:255],
            file_size_bytes=len(raw),
            description=desc,
            uploaded_by_user_id=uid,
            uploaded_by_name=uploader[:120],
        )
        db.session.add(proc)
        db.session.commit()
        auto_body = f"Novo processo: {name} {version} 📄"
        if len(auto_body) > 60:
            auto_body = auto_body[:57] + "…"
        plink = url_for("crm_news") + "#mf-process-" + str(proc.id)
        db.session.add(
            NewsPost(
                body=auto_body,
                category=_process_area_to_news_category(area),
                author_name=uploader[:80],
                author_role="Biblioteca",
                link_url=plink,
                process_area_tag=area[:40],
            )
        )
        db.session.commit()
        return jsonify(
            {
                "ok": True,
                "item": _serialize_process_library_item(proc),
                "summary": _process_library_api_summary_dict(),
            }
        )

    @app.route("/crm/process-library/<int:pid>/update", methods=["POST"])
    def crm_process_library_update(pid):
        proc = ProcessLibraryItem.query.get_or_404(pid)
        if not _can_manage_process_item(proc):
            return jsonify({"ok": False, "error": "Sem permissão para editar."}), 403
        name = (request.form.get("name") or proc.name).strip()[:200]
        area = (request.form.get("area") or proc.area).strip()
        if area not in PROCESS_LIBRARY_AREAS:
            area = proc.area
        proc.name = name
        proc.area = area
        proc.version = (request.form.get("version") or proc.version or "v1.0").strip()[:24]
        proc.responsible = (request.form.get("responsible") or proc.responsible or "")[:120]
        proc.description = (request.form.get("description") or "")[:120]
        f = request.files.get("file")
        if f and f.filename and f.filename.lower().endswith(".pdf"):
            raw = f.read()
            if len(raw) > 20 * 1024 * 1024:
                return jsonify({"ok": False, "error": "Tamanho máximo: 20 MB."}), 400
            if raw[:4] != b"%PDF":
                return jsonify({"ok": False, "error": "PDF inválido."}), 400
            base = current_app.config["PROCESS_LIBRARY_STORAGE"]
            old_fp = os.path.join(base, proc.stored_filename)
            if os.path.isfile(old_fp):
                try:
                    os.remove(old_fp)
                except OSError:
                    pass
            nfn = uuid.uuid4().hex + ".pdf"
            with open(os.path.join(base, nfn), "wb") as out:
                out.write(raw)
            proc.stored_filename = nfn
            proc.original_filename = (f.filename or "documento.pdf")[:255]
            proc.file_size_bytes = len(raw)
        db.session.commit()
        return jsonify(
            {"ok": True, "item": _serialize_process_library_item(proc), "summary": _process_library_api_summary_dict()}
        )

    @app.route("/crm/process-library/<int:pid>/delete", methods=["POST"])
    def crm_process_library_delete(pid):
        proc = ProcessLibraryItem.query.get_or_404(pid)
        if not _can_manage_process_item(proc):
            return jsonify({"ok": False, "error": "Sem permissão para excluir."}), 403
        base = current_app.config["PROCESS_LIBRARY_STORAGE"]
        fp = os.path.join(base, proc.stored_filename)
        if os.path.isfile(fp):
            try:
                os.remove(fp)
            except OSError:
                pass
        db.session.delete(proc)
        db.session.commit()
        return jsonify({"ok": True, "summary": _process_library_api_summary_dict()})

    @app.route("/crm/news/api/read", methods=["POST"])
    def crm_news_api_read():
        _news_session_user_key()
        data = request.get_json(silent=True) or {}
        pid = data.get("post_id")
        try:
            pid = int(pid)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "post_id"}), 400
        if not NewsPost.query.get(pid):
            return jsonify({"ok": False, "error": "not found"}), 404
        key = _news_session_user_key()
        if NewsReadReceipt.query.filter_by(post_id=pid, user_key=key).first():
            return jsonify({"ok": True})
        db.session.add(
            NewsReadReceipt(
                post_id=pid,
                user_key=key,
                reader_name=_news_reader_display_name(),
            )
        )
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
        return jsonify({"ok": True})

    @app.route("/crm/news/api/reads/<int:post_id>")
    def crm_news_api_reads(post_id):
        rows = (
            NewsReadReceipt.query.filter_by(post_id=post_id)
            .order_by(NewsReadReceipt.read_at.desc())
            .limit(200)
            .all()
        )
        return jsonify(
            {
                "reads": [
                    {
                        "name": r.reader_name,
                        "readAt": r.read_at.isoformat() + "Z" if r.read_at else "",
                    }
                    for r in rows
                ]
            }
        )

    @app.route("/crm/news/api/react", methods=["POST"])
    def crm_news_api_react():
        """Reações = status por post: ✅ Feito, ❌ Não feito, ⏳ Aguardando (um por utilizador)."""
        _news_session_user_key()
        data = request.get_json(silent=True) or {}
        try:
            pid = int(data.get("post_id"))
        except (TypeError, ValueError):
            return jsonify({"ok": False}), 400
        raw_emoji = (data.get("emoji") or "✅") or "✅"
        emoji = str(raw_emoji)[:16].strip() or "✅"
        # Apenas estes três estão disponíveis na UI; outros ignorados → ✅
        _status_set = frozenset({"✅", "❌", "⏳"})
        if emoji not in _status_set:
            emoji = "✅"
        if not NewsPost.query.get(pid):
            return jsonify({"ok": False}), 404
        key = _news_session_user_key()
        ex = NewsReaction.query.filter_by(post_id=pid, user_key=key, emoji=emoji).first()
        if ex:
            db.session.delete(ex)
        else:
            for other in _status_set:
                if other == emoji:
                    continue
                for old in NewsReaction.query.filter_by(post_id=pid, user_key=key, emoji=other).all():
                    db.session.delete(old)
            db.session.add(NewsReaction(post_id=pid, user_key=key, emoji=emoji))
        db.session.commit()
        reacts = (
            db.session.query(NewsReaction.emoji, db.func.count(NewsReaction.id))
            .filter(NewsReaction.post_id == pid)
            .group_by(NewsReaction.emoji)
            .all()
        )
        return jsonify(
            {
                "ok": True,
                "reactions": [{"emoji": e, "count": c} for e, c in reacts],
                "mine": [r.emoji for r in NewsReaction.query.filter_by(post_id=pid, user_key=key).all()],
            }
        )

    @app.route("/crm/linhas-telefonicas")
    @app.route("/crm/linhas_telefonicas")
    def crm_linhas_telefonicas():
        group_filter = request.args.get("group", "")
        search_q = request.args.get("q", "").strip()
        sort_col = request.args.get("sort", "number")
        order_dir = request.args.get("order", "asc")
        q = PhoneLine.query
        if group_filter:
            q = q.filter(PhoneLine.group_name == group_filter)
        if search_q:
            q = q.filter(db.or_(
                PhoneLine.number.ilike(f"%{search_q}%"),
                PhoneLine.name.ilike(f"%{search_q}%"),
                PhoneLine.group_name.ilike(f"%{search_q}%"),
            ))
        order_col = getattr(PhoneLine, sort_col, PhoneLine.number)
        if order_dir == "desc":
            q = q.order_by(order_col.desc())
        else:
            q = q.order_by(order_col.asc())
        lines = q.all()
        all_lines = PhoneLine.query.all()
        total = len(lines)
        total_all = len(all_lines)
        total_amount = sum((l.amount or 0) for l in lines)
        avg_amount = total_amount / total if total else 0
        groups = sorted(set(l.group_name for l in all_lines if l.group_name))
        with_wa = sum(1 for l in lines if l.has_whatsapp)
        without_wa = total - with_wa
        g_slot_counts = {f"G{i}": 0 for i in range(1, 13)}
        g_slot_counts["Outros"] = 0
        for l in all_lines:
            raw_g = (l.group_name or "").strip()
            m = re.match(r"^G(\d{1,2})$", raw_g, re.I)
            if m:
                n = int(m.group(1))
                if 1 <= n <= 12:
                    g_slot_counts[f"G{n}"] += 1
                else:
                    g_slot_counts["Outros"] += 1
            else:
                g_slot_counts["Outros"] += 1
        att_labels = list(g_slot_counts.keys())
        att_values = list(g_slot_counts.values())
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
        return render_template(
            "crm/linhas_telefonicas.html",
            lines=lines,
            total=total,
            total_all=total_all,
            total_amount=total_amount,
            avg_amount=avg_amount,
            groups=groups,
            with_whatsapp=with_wa,
            without_whatsapp=without_wa,
            last_updated=now,
            group_filter=group_filter,
            search_q=search_q,
            sort_col=sort_col,
            order_dir=order_dir,
            att_labels=att_labels,
            att_values=att_values,
        )

    @app.route("/crm/linhas-telefonicas/upload", methods=["POST"])
    def crm_linhas_telefonicas_upload():
        f = request.files.get("file")
        replace = request.form.get("replace_data") == "1"
        if not f or f.filename == "":
            flash("Selecione um arquivo CSV ou Excel.", "warning")
            return redirect(url_for("crm_linhas_telefonicas"))
        ext = (f.filename or "").lower().split(".")[-1]
        if ext not in ("csv", "xlsx"):
            flash("Formato inválido. Use .csv ou .xlsx.", "danger")
            return redirect(url_for("crm_linhas_telefonicas"))
        try:
            if replace:
                PhoneLine.query.delete()
                db.session.commit()
            rows = []
            def _parse_amount(v):
                if v is None or v == "": return 0.0
                s = str(v).strip().replace(",", "").replace(" ", "")
                s = re.sub(r"[^\d.\-]", "", s)
                try: return float(s) if s else 0.0
                except: return 0.0
            if ext == "csv":
                content = f.read().decode("utf-8-sig", errors="replace")
                reader = csv.DictReader(io.StringIO(content))
                for r in reader:
                    row = {k.strip().lower(): v for k, v in r.items() if k}
                    num = _normalize_source(row.get("number") or row.get("numero") or row.get("telefone") or row.get("phone") or "")
                    if not num:
                        continue
                    name = _normalize_source(row.get("name") or row.get("nome") or row.get("user") or row.get("usuario") or "")
                    group = _normalize_source(row.get("group") or row.get("grupo") or "")
                    amount = _parse_amount(row.get("amount") or row.get("total") or row.get("valor") or "")
                    wa = str(row.get("whatsapp") or row.get("wa") or row.get("has_whatsapp") or "").strip().lower() in ("1", "sim", "yes", "s", "y", "true")
                    rows.append(PhoneLine(number=num, name=name or None, group_name=group or None, amount=amount, has_whatsapp=wa))
            else:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip().lower() for c in ws[1]]
                for row in ws.iter_rows(min_row=2):
                    vals = {h: (row[i].value if i < len(row) else None) for i, h in enumerate(headers) if h}
                    num = _normalize_source(vals.get("number") or vals.get("numero") or vals.get("telefone") or vals.get("phone") or "")
                    if not num:
                        continue
                    name = _normalize_source(str(vals.get("name") or vals.get("nome") or vals.get("user") or "") or "")
                    group = _normalize_source(str(vals.get("group") or vals.get("grupo") or "") or "")
                    amount = _parse_amount(vals.get("amount") or vals.get("total") or vals.get("valor"))
                    wa_val = str(vals.get("whatsapp") or vals.get("wa") or vals.get("has_whatsapp") or "").strip().lower()
                    wa = wa_val in ("1", "sim", "yes", "s", "y", "true")
                    rows.append(PhoneLine(number=num, name=name or None, group_name=group or None, amount=amount, has_whatsapp=wa))
            for pl in rows:
                db.session.add(pl)
            db.session.commit()
            flash(f"{len(rows)} linha(s) importada(s) com sucesso.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao importar: {e}", "danger")
        return redirect(url_for("crm_linhas_telefonicas"))

    @app.route("/crm/linhas-telefonicas/download")
    def crm_linhas_telefonicas_download():
        lines = PhoneLine.query.order_by(PhoneLine.number).all()
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["number", "name", "group", "amount", "whatsapp"])
        for l in lines:
            w.writerow([l.number, l.name or "", l.group_name or "", l.amount or 0, "Sim" if l.has_whatsapp else "Não"])
        from flask import Response
        return Response(output.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": "attachment; filename=linhas_telefonicas.csv"})

    # --- Institutional modules: PW, Pool, Land Scape ---
    from website.crm_facility import parse_facility_csv, aggregate_charts

    _FACILITY_MODULES = {
        "pw": {
            "title": "PW",
            "donut_title": "Distribuição por tipo de PW",
            "bar_title": "Casas com maior volume de PW",
            "kpi_label": "Total de registros PW",
            "export_name": "pw_registros.csv",
        },
        "pool": {
            "title": "Pool",
            "donut_title": "Tipos de serviços de piscina",
            "bar_title": "Quantidade de casas por empresa",
            "kpi_label": "Total de serviços Pool",
            "export_name": "pool_servicos.csv",
        },
        "landscape": {
            "title": "Land Scape",
            "donut_title": "Tipos de serviços de paisagismo",
            "bar_title": "Casas com maior volume de Land Scape",
            "kpi_label": "Total de registros Land Scape",
            "export_name": "landscape_registros.csv",
        },
        "bbq": {
            "title": "BBQ",
            "donut_title": "Distribuição por tipo de serviço BBQ",
            "bar_title": "Casas com maior volume de BBQ",
            "kpi_label": "Total de registros BBQ",
            "export_name": "bbq_servicos.csv",
        },
    }

    def _facility_home_local_storage_payload(module_key: str):
        """Snapshot of CRM rows for browser localStorage (Home cards)."""
        if module_key not in ("pw", "pool", "bbq"):
            return None
        rows = (
            CrmServiceRecord.query.filter_by(module=module_key)
            .order_by(CrmServiceRecord.record_date.asc(), CrmServiceRecord.id.asc())
            .limit(20000)
            .all()
        )
        out = []
        for r in rows:
            out.append(
                {
                    "d": r.record_date.isoformat() if r.record_date else None,
                    "a": float(r.amount or 0),
                    "co": (r.category or "")[:120],
                    "prop": (r.property_name or "")[:120],
                    "typ": (r.service_type or "")[:120],
                }
            )
        return {"savedAt": datetime.utcnow().isoformat() + "Z", "rows": out}

    def _decode_csv_upload(raw_bytes):
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                return raw_bytes.decode(enc)
            except UnicodeDecodeError:
                continue
        return raw_bytes.decode("utf-8", errors="replace")

    def _facility_user_label():
        u = current_user()
        if u:
            return (getattr(u, "username", None) or getattr(u, "email", None) or str(u.id))[:120]
        return (get_client_ip() or "guest")[:120]

    def _facility_apply_filters(base_query, module_key, args):
        q = base_query.filter(CrmServiceRecord.module == module_key)
        month = (args.get("month") or "").strip()
        status = (args.get("status") or "").strip()
        category = (args.get("category") or "").strip()
        owner = (args.get("owner") or "").strip()
        tipo = (args.get("tipo") or "").strip()
        if month and len(month) >= 7:
            try:
                y, m = month.split("-", 1)
                yi, mi = int(y), int(m)
                from datetime import date as _date
                start = _date(yi, mi, 1)
                last_day = monthrange(yi, mi)[1]
                end = _date(yi, mi, last_day)
                q = q.filter(
                    CrmServiceRecord.record_date.isnot(None),
                    CrmServiceRecord.record_date >= start,
                    CrmServiceRecord.record_date <= end,
                )
            except (ValueError, TypeError):
                pass
        if status:
            q = q.filter(CrmServiceRecord.status == status)
        if category:
            q = q.filter(CrmServiceRecord.category == category)
        if owner:
            q = q.filter(CrmServiceRecord.owner.ilike(f"%{owner}%"))
        if tipo:
            q = q.filter(CrmServiceRecord.service_type == tipo)
        return q

    def _facility_distinct_options(module_key):
        bf = CrmServiceRecord.module == module_key
        dates = (
            db.session.query(CrmServiceRecord.record_date)
            .filter(bf, CrmServiceRecord.record_date.isnot(None))
            .distinct()
            .all()
        )
        months = sorted({d[0].strftime("%Y-%m") for d in dates if d[0]}, reverse=True)
        statuses = sorted(
            {s[0] for s in db.session.query(CrmServiceRecord.status).filter(bf).distinct() if s[0]}
        )
        categories = sorted(
            {
                c[0]
                for c in db.session.query(CrmServiceRecord.category).filter(bf).distinct()
                if c[0] and c[0] != "—"
            }
        )
        owners = sorted(
            {
                o[0]
                for o in db.session.query(CrmServiceRecord.owner).filter(bf).distinct()
                if o[0] and o[0] != "—"
            }
        )[:80]
        tipos = sorted(
            {
                t[0]
                for t in db.session.query(CrmServiceRecord.service_type).filter(bf).distinct()
                if t[0] and t[0] != "—"
            }
        )
        return months, statuses, categories, owners, tipos

    def _facility_render(module_key, cfg):
        def _run_pool_xlsx_import(f, raw_bytes, replace_records: bool):
            from website.crm_pool_dashboard import parse_pool_dashboard_xlsx, pool_rows_to_crm_records

            user_lbl = _facility_user_label()
            prev_count = CrmServiceRecord.query.filter_by(module="pool").count()
            payload, errs, m_rows, e_rows = parse_pool_dashboard_xlsx(raw_bytes, f.filename)
            if errs and not payload:
                try:
                    db.session.add(
                        CrmModuleUploadLog(
                            module="pool",
                            action="xlsx_dashboard",
                            filename=f.filename[:250],
                            rows_imported=0,
                            validation_ok=False,
                            message="; ".join(errs)[:500],
                            user_label=user_lbl,
                        )
                    )
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                flash("; ".join(errs[:8]), "danger")
                return
            if not payload:
                flash("Não foi possível ler dados das abas Monthly/Extra.", "danger")
                return
            try:
                period_key = ""
                py, pm = payload.get("period_year"), payload.get("period_month")
                if py is not None and pm is not None:
                    try:
                        period_key = f"{int(py)}-{int(pm):02d}"
                    except (TypeError, ValueError):
                        period_key = ""
                db.session.add(
                    PoolDashboardSnapshot(
                        period_key=period_key or "unknown",
                        payload_json=json.dumps(payload, ensure_ascii=False),
                        confirmation=(payload.get("confirmation") or "")[:500],
                        filename=f.filename[:250],
                    )
                )
                if replace_records:
                    CrmServiceRecord.query.filter_by(module="pool").delete()
                n_new = 0
                for row in pool_rows_to_crm_records(m_rows, e_rows):
                    db.session.add(CrmServiceRecord(**row))
                    n_new += 1
                db.session.add(
                    CrmModuleUploadLog(
                        module="pool",
                        action="xlsx_dashboard",
                        filename=f.filename[:250],
                        rows_imported=len(m_rows) + len(e_rows),
                        validation_ok=True,
                        message=(payload.get("confirmation") or "OK")[:500],
                        user_label=user_lbl,
                    )
                )
                db.session.commit()
                if replace_records:
                    flash(
                        f"✓ Substituído — {n_new} registros. Anterior removido.",
                        "success",
                    )
                else:
                    flash(
                        f"✓ Integrado — {n_new} novos + {prev_count} existentes = {prev_count + n_new}",
                        "success",
                    )
                if errs:
                    flash("Avisos: " + "; ".join(errs[:6]), "warning")
            except Exception as e:
                db.session.rollback()
                flash(f"Erro ao gravar dashboard: {e}", "danger")

        if (
            module_key == "pool"
            and request.method == "POST"
            and request.form.get("_action") == "upload_pool_unified"
        ):
            f = request.files.get("pool_unified_file")
            if not f or not f.filename:
                flash("Selecione um arquivo Excel ou CSV.", "warning")
            else:
                low = f.filename.lower()
                raw = f.read()
                if low.endswith(".xlsx"):
                    mode = (request.form.get("pool_xlsx_mode") or "replace").strip().lower()
                    replace_x = mode != "append"
                    _run_pool_xlsx_import(f, raw, replace_x)
                elif low.endswith(".csv"):
                    text = _decode_csv_upload(raw)
                    rows, errs, mapping = parse_facility_csv(text, "pool")
                    user_lbl = _facility_user_label()
                    upload_mode = (request.form.get("upload_mode") or "append").strip().lower()
                    replace = upload_mode == "replace"
                    prev_count = CrmServiceRecord.query.filter_by(module="pool").count()
                    if errs:
                        db.session.add(
                            CrmModuleUploadLog(
                                module="pool",
                                action=upload_mode,
                                filename=f.filename[:250],
                                rows_imported=0,
                                validation_ok=False,
                                message="; ".join(errs)[:500],
                                user_label=user_lbl,
                            )
                        )
                        db.session.commit()
                        flash("Validação falhou: " + "; ".join(errs[:5]), "danger")
                    else:
                        try:
                            if replace:
                                CrmServiceRecord.query.filter_by(module="pool").delete()
                            for r in rows:
                                db.session.add(CrmServiceRecord(**r))
                            db.session.add(
                                CrmModuleUploadLog(
                                    module="pool",
                                    action=upload_mode,
                                    filename=f.filename[:250],
                                    rows_imported=len(rows),
                                    validation_ok=True,
                                    message="OK",
                                    user_label=user_lbl,
                                )
                            )
                            db.session.commit()
                            if replace:
                                flash(
                                    f"✓ Substituído — {len(rows)} registros. Anterior removido.",
                                    "success",
                                )
                            else:
                                flash(
                                    f"✓ Integrado — {len(rows)} novos + {prev_count} existentes = {prev_count + len(rows)}",
                                    "success",
                                )
                        except Exception as e:
                            db.session.rollback()
                            db.session.add(
                                CrmModuleUploadLog(
                                    module="pool",
                                    action=upload_mode,
                                    filename=f.filename[:250],
                                    rows_imported=0,
                                    validation_ok=False,
                                    message=str(e)[:500],
                                    user_label=user_lbl,
                                )
                            )
                            db.session.commit()
                            flash(f"Erro ao gravar: {e}", "danger")
                else:
                    flash("Use .xlsx (Dashboard_Piscineiros…) ou .csv legado.", "warning")
            return redirect(url_for("crm_pool", **request.args.to_dict(flat=True)))

        if (
            module_key == "pool"
            and request.method == "POST"
            and request.form.get("_action") == "upload_pool_dashboard"
        ):
            f = request.files.get("pool_dashboard_xlsx")
            user_lbl = _facility_user_label()
            if not f or not f.filename:
                flash("Selecione a planilha Excel (.xlsx).", "warning")
            elif not f.filename.lower().endswith(".xlsx"):
                flash(
                    "Use arquivo .xlsx (ex.: Dashboard_Piscineiros_x_Enderecos_March_2025.xlsx).",
                    "warning",
                )
            else:
                raw = f.read()
                mode = (request.form.get("pool_xlsx_mode") or "replace").strip().lower()
                replace_x = mode != "append"
                _run_pool_xlsx_import(f, raw, replace_x)
            return redirect(url_for("crm_pool", **request.args.to_dict(flat=True)))

        if request.method == "POST" and request.form.get("_action") == "upload":
            f = request.files.get("csv_file") or request.files.get("file")
            upload_mode = (request.form.get("upload_mode") or "append").strip().lower()
            replace = upload_mode == "replace"
            if not f or not f.filename:
                flash("Selecione um arquivo CSV.", "warning")
            elif not f.filename.lower().endswith(".csv"):
                flash("Use arquivo .csv", "warning")
            else:
                raw = f.read()
                text = _decode_csv_upload(raw)
                rows, errs, mapping = parse_facility_csv(text, module_key)
                user_lbl = _facility_user_label()
                if errs:
                    db.session.add(
                        CrmModuleUploadLog(
                            module=module_key,
                            action=upload_mode,
                            filename=f.filename[:250],
                            rows_imported=0,
                            validation_ok=False,
                            message="; ".join(errs)[:500],
                            user_label=user_lbl,
                        )
                    )
                    db.session.commit()
                    flash("Validação falhou: " + "; ".join(errs[:5]), "danger")
                else:
                    try:
                        prev_count = CrmServiceRecord.query.filter_by(module=module_key).count()
                        if replace:
                            CrmServiceRecord.query.filter_by(module=module_key).delete()
                        for r in rows:
                            db.session.add(CrmServiceRecord(**r))
                        db.session.add(
                            CrmModuleUploadLog(
                                module=module_key,
                                action=upload_mode,
                                filename=f.filename[:250],
                                rows_imported=len(rows),
                                validation_ok=True,
                                message="OK",
                                user_label=user_lbl,
                            )
                        )
                        db.session.commit()
                        if replace:
                            flash(
                                f"✓ Substituído — {len(rows)} registros. Anterior removido.",
                                "success",
                            )
                        else:
                            flash(
                                f"✓ Integrado — {len(rows)} novos + {prev_count} existentes = {prev_count + len(rows)}",
                                "success",
                            )
                    except Exception as e:
                        db.session.rollback()
                        db.session.add(
                            CrmModuleUploadLog(
                                module=module_key,
                                action=upload_mode,
                                filename=f.filename[:250],
                                rows_imported=0,
                                validation_ok=False,
                                message=str(e)[:500],
                                user_label=user_lbl,
                            )
                        )
                        db.session.commit()
                        flash(f"Erro ao gravar: {e}", "danger")
            return redirect(url_for(f"crm_{module_key}", **request.args.to_dict(flat=True)))

        # View log (GET)
        try:
            db.session.add(
                CrmModuleViewLog(
                    module=module_key,
                    user_label=_facility_user_label(),
                    ip_address=(get_client_ip() or "")[:64],
                )
            )
            db.session.commit()
        except Exception:
            db.session.rollback()

        q = _facility_apply_filters(CrmServiceRecord.query, module_key, request.args)
        records = q.order_by(
            CrmServiceRecord.record_date.is_(None),
            CrmServiceRecord.record_date.desc(),
            CrmServiceRecord.id.desc(),
        ).all()
        agg = aggregate_charts(records, module_key=module_key)
        months, statuses, categories, owners, tipos = _facility_distinct_options(module_key)
        upload_logs = (
            CrmModuleUploadLog.query.filter_by(module=module_key)
            .order_by(CrmModuleUploadLog.created_at.desc())
            .limit(40)
            .all()
        )
        view_logs = (
            CrmModuleViewLog.query.filter_by(module=module_key)
            .order_by(CrmModuleViewLog.created_at.desc())
            .limit(40)
            .all()
        )
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
        pool_dashboard = None
        pool_history_chart = None
        pool_snapshot_exists = 0
        facility_total_records = CrmServiceRecord.query.filter_by(module=module_key).count()
        if module_key == "pool":
            _snap = (
                PoolDashboardSnapshot.query.order_by(PoolDashboardSnapshot.created_at.desc())
                .first()
            )
            if _snap:
                pool_snapshot_exists = 1
            if _snap and _snap.payload_json:
                try:
                    pool_dashboard = json.loads(_snap.payload_json)
                except (json.JSONDecodeError, TypeError):
                    pool_dashboard = None
            cutoff = datetime.utcnow() - timedelta(days=30)
            logs30 = (
                CrmModuleUploadLog.query.filter(
                    CrmModuleUploadLog.module == "pool",
                    CrmModuleUploadLog.created_at >= cutoff,
                )
                .order_by(CrmModuleUploadLog.created_at.asc())
                .all()
            )
            by_day = defaultdict(lambda: {"imports": 0, "rows": 0})
            for log in logs30:
                if not log.created_at:
                    continue
                d = log.created_at.strftime("%Y-%m-%d")
                by_day[d]["imports"] += 1
                by_day[d]["rows"] += int(log.rows_imported or 0)
            labels = sorted(by_day.keys())
            pool_history_chart = {
                "labels": labels,
                "imports": [by_day[d]["imports"] for d in labels],
                "rows": [by_day[d]["rows"] for d in labels],
            }
        facility_home_sync = _facility_home_local_storage_payload(module_key)
        return render_template(
            "crm/service_module_institutional.html",
            facility_cfg=cfg,
            module_key=module_key,
            records=records[:500],
            agg=agg,
            filter_months=months,
            filter_statuses=statuses,
            filter_categories=categories,
            filter_owners=owners,
            filter_tipos=tipos,
            upload_logs=upload_logs,
            view_logs=view_logs,
            last_updated=now,
            pool_dashboard=pool_dashboard,
            pool_history_chart=pool_history_chart,
            facility_total_records=facility_total_records,
            pool_snapshot_exists=pool_snapshot_exists,
            facility_home_sync=facility_home_sync,
            current_filters={
                "month": request.args.get("month", ""),
                "status": request.args.get("status", ""),
                "category": request.args.get("category", ""),
                "owner": request.args.get("owner", ""),
                "tipo": request.args.get("tipo", ""),
            },
        )

    def _facility_export(module_key, cfg):
        q = _facility_apply_filters(CrmServiceRecord.query, module_key, request.args)
        rows = q.order_by(CrmServiceRecord.id).all()
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(
            ["date", "property", "category", "owner", "status", "type", "amount", "created_at"]
        )
        for r in rows:
            w.writerow(
                [
                    r.record_date.isoformat() if r.record_date else "",
                    r.property_name or "",
                    r.category or "",
                    r.owner or "",
                    r.status or "",
                    r.service_type or "",
                    r.amount or 0,
                    r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
                ]
            )
        return Response(
            output.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={cfg['export_name']}"},
        )

    @app.route("/crm/pw", methods=["GET", "POST"])
    def crm_pw():
        return _facility_render("pw", _FACILITY_MODULES["pw"])

    @app.route("/crm/pw/export")
    def crm_pw_export():
        return _facility_export("pw", _FACILITY_MODULES["pw"])

    @app.route("/crm/pool", methods=["GET", "POST"])
    def crm_pool():
        return _facility_render("pool", _FACILITY_MODULES["pool"])

    @app.route("/crm/pool/export")
    def crm_pool_export():
        return _facility_export("pool", _FACILITY_MODULES["pool"])

    @app.route("/crm/landscape", methods=["GET", "POST"])
    @app.route("/crm/land-scape", methods=["GET", "POST"])
    def crm_landscape():
        return _facility_render("landscape", _FACILITY_MODULES["landscape"])

    @app.route("/crm/landscape/export")
    @app.route("/crm/land-scape/export")
    def crm_landscape_export():
        return _facility_export("landscape", _FACILITY_MODULES["landscape"])

    @app.route("/crm/bbq", methods=["GET", "POST"])
    def crm_bbq():
        return _facility_render("bbq", _FACILITY_MODULES["bbq"])

    @app.route("/crm/bbq/export")
    def crm_bbq_export():
        return _facility_export("bbq", _FACILITY_MODULES["bbq"])


def register_routes(app):
    """Register all application routes."""

    # CRM routes first so crm_home exists for redirects
    _register_crm_routes(app)

    from website.qb_routes import register_quickbooks_routes

    register_quickbooks_routes(app)

    @app.route("/set-language")
    def set_language():
        lang = request.args.get("lang", "en").lower()
        if lang not in ("en", "pt", "es"):
            lang = "en"
        session["locale"] = lang
        next_url = request.referrer or url_for("crm_home")
        return redirect(next_url)

    @app.context_processor
    def inject_locale():
        from website.translations import get_locale, t, TRANSLATIONS
        locale = get_locale()
        return {"locale": locale, "t": t, "TRANSLATIONS": TRANSLATIONS}

    @app.route("/")
    def index():
        return redirect(url_for("crm_home"))
    
    @app.route("/login", methods=["GET", "POST"])
    def login():
        return redirect(url_for("crm_home"))
    
    @app.route("/logout")
    def logout():
        return redirect(url_for("crm_home"))
    
    @app.route("/check_username", methods=["GET"])
    def check_username():
        """AJAX endpoint to check username availability."""
        username = request.args.get("username", "").strip().lower()
        if not username:
            return jsonify({"available": False})
        
        exists = User.query.filter_by(username=username).first() is not None
        return jsonify({"available": not exists})
    
    @app.route("/request_account", methods=["POST"])
    def request_account():
        """Handle account request submission."""
        username = request.form.get("username", "").strip().lower()
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        
        # Validation
        if not username or not first_name or not last_name or not email or not password:
            flash("All fields are required.", "danger")
            return redirect(url_for("crm_home"))
        
        if password != confirm_password:
            flash("Passwords do not match.", "danger")
            return redirect(url_for("crm_home"))
        
        if not is_password_strong(password):
            flash("Password does not meet strength requirements.", "danger")
            return redirect(url_for("crm_home"))
        
        if User.query.filter_by(username=username).first():
            flash("Username already exists.", "danger")
            return redirect(url_for("crm_home"))
        
        if User.query.filter_by(email=email).first():
            flash("Email already registered.", "danger")
            return redirect(url_for("crm_home"))
        
        # Create account request
        request_obj = AccountRequest(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            password_hash=hash_password(password),
            ip_address=get_client_ip(),
            user_agent=get_user_agent(),
            status="pending"
        )
        
        db.session.add(request_obj)
        db.session.commit()
        
        flash("Account request submitted successfully. An admin will review your request.", "success")
        return redirect(url_for("crm_home"))
    
    @app.route("/dashboard")
    def dashboard():
        user = current_user()
        if not user:
            return redirect(url_for("crm_home"))
        return render_template("dashboard.html", user=user)

    @app.route("/admin")
    @admin_required
    def admin():
        user = current_user()
        pending_requests = AccountRequest.query.filter_by(status="pending").order_by(AccountRequest.created_at.desc()).all()
        users = User.query.order_by(User.created_at.desc()).all()
        return render_template("admin.html", user=user, pending=pending_requests, users=users)
    
    @app.route("/admin/users")
    @admin_required
    def admin_users():
        user = current_user()
        users = User.query.order_by(User.created_at.desc()).all()
        return render_template("admin_users.html", user=user, users=users)
    
    @app.route("/admin/approve_request/<int:request_id>", methods=["POST"])
    @admin_required
    def approve_request(request_id):
        request_obj = AccountRequest.query.get_or_404(request_id)
        if request_obj.status != "pending":
            flash("Request already processed.", "warning")
            return redirect(url_for("admin"))
        
        # Check if username already exists
        if User.query.filter_by(username=request_obj.username).first():
            flash("Username already exists. Cannot approve request.", "danger")
            request_obj.status = "rejected"
            db.session.commit()
            return redirect(url_for("admin"))
        
        # Create user account
        new_user = User(
            username=request_obj.username,
            email=request_obj.email,
            password_hash=request_obj.password_hash,
            first_name=request_obj.first_name,
            last_name=request_obj.last_name,
            is_admin=False,
            is_active=True,
            email_verified=True
        )
        
        db.session.add(new_user)
        request_obj.status = "approved"
        request_obj.processed_by = current_user().id
        request_obj.processed_at = datetime.utcnow()
        db.session.commit()
        
        flash("Account request approved successfully.", "success")
        return redirect(url_for("admin"))
    
    @app.route("/admin/reject_request/<int:request_id>", methods=["POST"])
    @admin_required
    def reject_request(request_id):
        request_obj = AccountRequest.query.get_or_404(request_id)
        if request_obj.status != "pending":
            flash("Request already processed.", "warning")
            return redirect(url_for("admin"))
        
        request_obj.status = "rejected"
        request_obj.processed_by = current_user().id
        request_obj.processed_at = datetime.utcnow()
        db.session.commit()
        
        flash("Account request rejected.", "info")
        return redirect(url_for("admin"))
    
    @app.route("/admin/create_user", methods=["POST"])
    @admin_required
    def create_user():
        username = request.form.get("username", "").strip().lower()
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        is_admin = bool(request.form.get("is_admin"))
        
        # Validation
        if not username or not email or not password:
            flash("Username, email, and password are required.", "danger")
            return redirect(url_for("admin_users"))
        
        if not is_password_strong(password):
            flash("Password does not meet strength requirements.", "danger")
            return redirect(url_for("admin_users"))
        
        if User.query.filter_by(username=username).first():
            flash("Username already exists.", "danger")
            return redirect(url_for("admin_users"))
        
        if User.query.filter_by(email=email).first():
            flash("Email already registered.", "danger")
            return redirect(url_for("admin_users"))
        
        # Create user
        new_user = User(
            username=username,
            email=email,
            password_hash=hash_password(password),
            first_name=first_name,
            last_name=last_name,
            is_admin=is_admin,
            is_active=True,
            email_verified=True
        )
        
        db.session.add(new_user)
        db.session.commit()
        
        flash("User created successfully.", "success")
        return redirect(url_for("admin_users"))
    
    @app.route("/admin/update_user/<int:user_id>", methods=["POST"])
    @admin_required
    def update_user(user_id):
        user = User.query.get_or_404(user_id)
        
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()
        is_admin = bool(request.form.get("is_admin"))
        is_active = bool(request.form.get("is_active"))
        
        # Update fields (username is read-only)
        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        user.is_admin = is_admin
        user.is_active = is_active
        
        # Update password if provided
        if password:
            if not is_password_strong(password):
                flash("Password does not meet strength requirements.", "danger")
                return redirect(url_for("admin_users"))
            user.password_hash = hash_password(password)
        
        # Check email uniqueness
        existing = User.query.filter_by(email=email).first()
        if existing and existing.id != user.id:
            flash("Email already registered to another user.", "danger")
            return redirect(url_for("admin_users"))
        
        db.session.commit()
        flash("User updated successfully.", "success")
        return redirect(url_for("admin_users"))
    
    @app.route("/admin/delete_user/<int:user_id>", methods=["POST"])
    @admin_required
    def delete_user(user_id):
        user = User.query.get_or_404(user_id)
        
        # Prevent deleting yourself
        if user.id == current_user().id:
            flash("You cannot delete your own account.", "danger")
            return redirect(url_for("admin_users"))
        
        db.session.delete(user)
        db.session.commit()
        
        flash("User deleted successfully.", "success")
        return redirect(url_for("admin_users"))


app = create_app()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)

